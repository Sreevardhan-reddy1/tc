#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║   ServiceNow Ticket Counter  —  Single-File Application      ║
║   Upload: RITM → Incident → MACM → Dashboard + Excel report  ║
╚══════════════════════════════════════════════════════════════╝
Run:  python app.py
URL:  http://localhost:5000
"""

# ═══════════════════════════════════════════════════════════════
# 0. AUTO-INSTALL DEPENDENCIES
# ═══════════════════════════════════════════════════════════════
import subprocess, sys, os

_PACKAGES = [
    "flask>=3.1.0", "pandas>=2.2.3", "openpyxl>=3.1.3",
    "pdfplumber>=0.11.4", "Pillow>=11.0.0",
    "werkzeug>=3.1.0", "xlrd>=2.0.1", "numpy>=2.1.0",
    "anthropic>=0.30.0",
]

def _install_packages():
    print("Checking / installing dependencies…")
    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", "--prefer-binary", "--quiet"] + _PACKAGES,
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print("Bulk install failed, trying individually…")
        for pkg in _PACKAGES:
            subprocess.run(
                [sys.executable, "-m", "pip", "install", "--prefer-binary", "--quiet", pkg],
                capture_output=True
            )
    print("Dependencies ready.\n")

# On cloud platforms packages are already installed via requirements.txt;
# running pip at startup over the network makes gunicorn miss health checks → 403.
_on_cloud = bool(os.environ.get("RENDER") or os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("DYNO"))
if not _on_cloud:
    _install_packages()

# ═══════════════════════════════════════════════════════════════
# 1. IMPORTS
# ═══════════════════════════════════════════════════════════════
import json, uuid, shutil, re, tempfile, threading, webbrowser
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from pathlib import Path

from flask import (Flask, render_template_string, request,
                   jsonify, session, send_file, redirect, url_for)
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    from PIL import Image
    import tesserocr
    TESSERACT_OK = True
except ImportError:
    TESSERACT_OK = False

# Windows 10/11 built-in OCR via PowerShell — no compilation or downloads needed
POWERSHELL_OK = sys.platform == "win32"
WINSDK_OK     = False   # legacy flag, kept for compatibility

EASYOCR_OK = False
_easyocr_reader = None

try:
    import anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False

OCR_OK = POWERSHELL_OK or TESSERACT_OK or EASYOCR_OK or ANTHROPIC_OK

# ═══════════════════════════════════════════════════════════════
# 2. APP CONFIGURATION
# ═══════════════════════════════════════════════════════════════
BASE_DIR        = Path(__file__).parent

# On cloud platforms (Render, Railway, Heroku) use /tmp which is
# always writable; fall back to the project directory locally.
_is_cloud = bool(
    os.environ.get("RENDER") or
    os.environ.get("RAILWAY_ENVIRONMENT") or
    os.environ.get("DYNO")
)
_DATA_ROOT      = Path("/tmp") if _is_cloud else BASE_DIR

UPLOAD_FOLDER   = _DATA_ROOT / "uploads"
OUTPUT_FOLDER   = _DATA_ROOT / "output"
REFERENCE_FOLDER= _DATA_ROOT / "reference_template"

for _d in [UPLOAD_FOLDER, OUTPUT_FOLDER, REFERENCE_FOLDER]:
    _d.mkdir(parents=True, exist_ok=True)

ALLOWED_EXT = {"xlsx","xls","csv","pdf","png","jpg","jpeg","bmp","tiff","gif"}
MONTH_ABBR  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_FULL  = ["January","February","March","April","May","June",
               "July","August","September","October","November","December"]

CHATBOT_CONFIG_FILE = BASE_DIR / "chatbot_config.json"
DEFAULT_CHATBOT_CONFIG = {
    "api_key": "",
    "model":   "claude-haiku-4-5-20251001",
}

def load_chatbot_config():
    if CHATBOT_CONFIG_FILE.exists():
        try:
            with open(CHATBOT_CONFIG_FILE) as f:
                cfg = json.load(f)
            merged = dict(DEFAULT_CHATBOT_CONFIG); merged.update(cfg)
            return merged
        except Exception:
            pass
    return dict(DEFAULT_CHATBOT_CONFIG)

def save_chatbot_config(cfg):
    with open(CHATBOT_CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

EMAIL_CONFIG_FILE = BASE_DIR / "email_config.json"
DEFAULT_EMAIL_CONFIG = {
    "smtp_host":     "smtp.gmail.com",
    "smtp_port":     587,
    "smtp_user":     "",
    "smtp_password": "",
    "sender_name":   "ServiceNow Ticket Counter",
    "recipients":    ["sreevardhanr7@gmail.com"],
    "use_tls":       True
}

app = Flask(__name__)
# Fix for reverse-proxy deployments (Render, Railway, Heroku):
# reads X-Forwarded-For / X-Forwarded-Proto headers so Flask
# generates correct HTTPS URLs and session cookies work properly.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

app.secret_key = os.environ.get("SECRET_KEY", "sn-ticket-counter-key-2024")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024   # 100 MB
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"]   = _is_cloud       # HTTPS-only on cloud


# ── Always return JSON for common HTTP errors ─────────────────
@app.errorhandler(404)
def err_not_found(_):
    return jsonify({"success": False, "error": "Not found (404)"}), 404

@app.errorhandler(403)
def err_forbidden(_):
    return jsonify({"success": False, "error": "Forbidden (403) — server blocked the request"}), 403

@app.errorhandler(413)
def err_too_large(_):
    return jsonify({"success": False, "error": "File too large (max 100 MB)"}), 413

@app.errorhandler(500)
def err_server(e):
    return jsonify({"success": False, "error": f"Internal server error: {e}"}), 500

# ═══════════════════════════════════════════════════════════════
# 3. COLUMN AUTO-DETECTION  (ServiceNow export column aliases)
# ═══════════════════════════════════════════════════════════════
_COL_ALIASES = {
    "number"  : ["number","ticket number","ticket #","ticket no","incident number",
                 "ritm number","req item","request item","macm number","task number","id"],
    "team"    : ["assignment group","assignmentgroup","assignment_group","team","group",
                 "support group","resolver group","assigned group","work group"],
    "assignee": ["assigned to","assigned_to","assignedto","assignee","resolved by",
                 "worked by","agent","technician","handled by","owner"],
    "state"   : ["state","status","ticket state","ticket status","current state"],
    "opened"  : ["opened","opened at","opened_at","open date",
                 "start","start date","start_date",
                 "created","created on","creation date",
                 "reported date","date opened","sys_created_on"],
    "closed"  : ["closed","closed at","closed_at","close date","resolved","resolved at",
                 "resolution date","end date","date closed","completed"],
    "short_description": ["short description","short_description","summary","description","title"],
    "priority": ["priority","severity","impact","urgency"],
    "month"   : ["month","period","reporting month","report month"],
}

def _norm(s): return str(s).strip().lower().replace("_"," ").replace("-"," ")

def _detect_col(columns, field):
    aliases = _COL_ALIASES.get(field, [field])
    norm_map = {_norm(c): c for c in columns}
    for a in aliases:
        if a in norm_map: return norm_map[a]
    for a in aliases:
        for n, orig in norm_map.items():
            if a in n or n in a: return orig
    return None

def _parse_dates(series):
    fmts = ["%Y-%m-%d %H:%M:%S","%Y-%m-%d","%d/%m/%Y %H:%M:%S","%d/%m/%Y",
            "%m/%d/%Y %H:%M:%S","%m/%d/%Y","%d-%m-%Y","%d %b %Y","%B %d, %Y"]
    for fmt in fmts:
        try:
            p = pd.to_datetime(series, format=fmt, errors="coerce")
            if p.notna().sum() > len(series) * 0.3: return p
        except Exception: pass
    return pd.to_datetime(series, errors="coerce", dayfirst=True)

# ═══════════════════════════════════════════════════════════════
# 4. EXCEL / CSV PROCESSOR
# ═══════════════════════════════════════════════════════════════
def process_excel(filepath, ticket_type):
    result = dict(ticket_type=ticket_type, total=0, columns_found={},
                  by_team={}, by_assignee={}, by_month={}, by_team_assignee={},
                  by_month_team={}, by_month_assignee={},
                  records=[], errors=[])
    try:
        if filepath.lower().endswith(".csv"):
            df = pd.read_csv(filepath, dtype=str, encoding="utf-8", errors="replace")
        else:
            xl = pd.ExcelFile(filepath)
            best = None
            for sh in xl.sheet_names:
                try:
                    tmp = pd.read_excel(filepath, sheet_name=sh, dtype=str)
                    if best is None or len(tmp) > len(best): best = tmp
                except Exception: pass
            if best is None:
                result["errors"].append("Could not read any sheet."); return result
            df = best

        if df.empty:
            result["errors"].append("File is empty."); return result

        df.columns = [str(c).strip() for c in df.columns]
        col_map = {f: _detect_col(list(df.columns), f) for f in _COL_ALIASES}
        col_map = {k: v for k, v in col_map.items() if v}
        result["columns_found"] = col_map

        ttype_upper = ticket_type.upper()

        # ── Date field: RITM/Incident → Start Date; MACM → Closed ─
        if ttype_upper in ("RITM", "INCIDENT"):
            for col in df.columns:
                if _norm(col) in ("start", "start date", "start_date",
                                  "scheduled start", "scheduled_start", "planned start"):
                    col_map["opened"] = col
                    break
            dk = "opened" if "opened" in col_map else None
        else:
            dk = "closed" if "closed" in col_map else ("opened" if "opened" in col_map else None)

        # ── Basic cleanup ─────────────────────────────────────────
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)
        if "number" in col_map:
            df[col_map["number"]] = df[col_map["number"]].astype(str).str.strip()
            df = df[df[col_map["number"]].str.len() > 2]
        df.reset_index(drop=True, inplace=True)

        # ── Pre-clean team / assignee ─────────────────────────────
        if "team" in col_map:
            df[col_map["team"]] = df[col_map["team"]].fillna("Unknown").astype(str).str.strip()
        if "assignee" in col_map:
            df[col_map["assignee"]] = df[col_map["assignee"]].fillna("Unassigned").astype(str).str.strip()

        # ── Save ALL records (pre-dedup) for the duplicates report ─
        df_all = df.copy()

        # ── Deduplicate by ticket NUMBER (same ticket exported multiple times) ──
        # Dedup by description was wrong: many legitimate tickets share the same
        # short description (e.g. "Password Reset" or "Server Down").
        if "number" in col_map:
            num_col = col_map["number"]
            dup_mask = df[num_col].duplicated(keep="first")
            dup_count = int(dup_mask.sum())
            if dup_count:
                result["duplicate_count"] = dup_count
            df = df[~dup_mask].reset_index(drop=True)

        # ── Total = unique tickets (all rows, date not required) ───
        result["total"] = len(df)

        if "team" in col_map:
            result["by_team"] = df[col_map["team"]].value_counts().to_dict()
        else:
            result["by_team"] = {"Unassigned": result["total"]}

        if "assignee" in col_map:
            result["by_assignee"] = df[col_map["assignee"]].value_counts().to_dict()

        if "team" in col_map and "assignee" in col_map:
            tc, ac = col_map["team"], col_map["assignee"]
            ta = {}
            for _, row in df.groupby([tc, ac]).size().reset_index(name="cnt").iterrows():
                ta.setdefault(str(row[tc]), {})[str(row[ac])] = int(row["cnt"])
            result["by_team_assignee"] = ta

        # ── Monthly breakdown: dated records only ──────────────────
        if dk and dk in col_map:
            dates = _parse_dates(df[col_map[dk]])
            valid = dates.notna()
            no_date = int((~valid).sum())
            if no_date:
                result["no_date_count"] = no_date
            dfd = df[valid].reset_index(drop=True)
            dfd["_month"] = dates[valid].reset_index(drop=True).dt.to_period("M").astype(str)
            result["by_month"] = dfd["_month"].value_counts().sort_index().to_dict()

            if "team" in col_map:
                tc = col_map["team"]
                mbt = {}
                for _, row in dfd.groupby(["_month", tc]).size().reset_index(name="cnt").iterrows():
                    mbt.setdefault(str(row["_month"]), {})[str(row[tc])] = int(row["cnt"])
                result["by_month_team"] = mbt

            if "assignee" in col_map:
                ac = col_map["assignee"]
                mba = {}
                for _, row in dfd.groupby(["_month", ac]).size().reset_index(name="cnt").iterrows():
                    mba.setdefault(str(row["_month"]), {})[str(row[ac])] = int(row["cnt"])
                result["by_month_assignee"] = mba

        elif "month" in col_map:
            result["by_month"] = df[col_map["month"]].fillna("Unknown").value_counts().sort_index().to_dict()

        # ── Store ALL records (incl. duplicates) for report ───────
        used_cols = list(dict.fromkeys(col_map.values()))
        canonical = [k for k, v in col_map.items()
                     if v in used_cols and
                     used_cols.index(v) == list(col_map.values()).index(v)]
        if used_cols:
            sub = df_all[used_cols].copy()
            sub.columns = canonical
            result["records"] = sub.fillna("").to_dict(orient="records")

        # ── Duration & Top-10 slowest tickets (RITM / Incident) ───
        if ttype_upper in ("RITM", "INCIDENT"):
            ocol = col_map.get("opened")
            ccol = col_map.get("closed")
            if ocol and ccol:
                try:
                    o_dates = _parse_dates(df[ocol])
                    c_dates = _parse_dates(df[ccol])
                    dur = (c_dates - o_dates).dt.days
                    df_dur = df.copy()
                    df_dur["_dur"] = dur
                    valid_dur = df_dur[df_dur["_dur"].notna() & (df_dur["_dur"] >= 0)]
                    if not valid_dur.empty:
                        top10 = valid_dur.nlargest(10, "_dur")
                        t10_rows = []
                        for _, row in top10.iterrows():
                            ncol  = col_map.get("number")
                            scol  = col_map.get("short_description")
                            tcol  = col_map.get("team")
                            acol  = col_map.get("assignee")
                            stcol = col_map.get("state")
                            pcol  = col_map.get("priority")
                            rec = {
                                "number":            str(row[ncol]).strip()  if ncol  and ncol  in row.index else "",
                                "short_description": str(row[scol]).strip()  if scol  and scol  in row.index else "",
                                "team":              str(row[tcol]).strip()  if tcol  and tcol  in row.index else "",
                                "assignee":          str(row[acol]).strip()  if acol  and acol  in row.index else "",
                                "state":             str(row[stcol]).strip() if stcol and stcol in row.index else "",
                                "priority":          str(row[pcol]).strip()  if pcol  and pcol  in row.index else "",
                                "opened":            str(row[ocol]).strip(),
                                "closed":            str(row[ccol]).strip(),
                                "duration_days":     int(row["_dur"]),
                                "ticket_type":       ttype_upper,
                            }
                            t10_rows.append(rec)
                        result["top10_slow"] = t10_rows
                except Exception:
                    pass

    except Exception as e:
        result["errors"].append(f"Excel error: {e}")
    return result

# ═══════════════════════════════════════════════════════════════
# 5. PDF PROCESSOR
# ═══════════════════════════════════════════════════════════════
def process_pdf(filepath, ticket_type):
    result = dict(ticket_type=ticket_type, total=0, columns_found={},
                  by_team={}, by_assignee={}, by_month={}, by_team_assignee={},
                  records=[], errors=[])
    if not PDFPLUMBER_OK:
        result["errors"].append("pdfplumber not installed."); return result
    try:
        all_rows, headers = [], None
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                for tbl in (page.extract_tables() or []):
                    if not tbl: continue
                    if headers is None and len(tbl) > 1:
                        headers = [str(c).strip() if c else f"Col_{i}" for i,c in enumerate(tbl[0])]
                        all_rows.extend(tbl[1:])
                    elif headers:
                        first = [str(c).strip() if c else "" for c in tbl[0]]
                        all_rows.extend(tbl[1:] if first == headers else tbl)

        if headers and all_rows:
            df = pd.DataFrame(all_rows, columns=headers[:len(all_rows[0])])
            df.dropna(how="all", inplace=True)
            tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False,
                                              mode="w", encoding="utf-8", newline="")
            df.to_csv(tmp.name, index=False); tmp.close()
            try:
                result = process_excel(tmp.name, ticket_type)
                result["ticket_type"] = ticket_type
            finally:
                os.unlink(tmp.name)
        else:
            with pdfplumber.open(filepath) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            nums = list(set(re.findall(r"(?:RITM|INC|MACM|TASK|REQ)\d+", text, re.IGNORECASE)))
            result["total"] = len(nums)
            result["records"] = [{"number": n} for n in nums]
            if nums:
                result["by_team"] = {"Extracted from PDF": len(nums)}
            else:
                result["errors"].append("No tables or ticket IDs found in PDF.")
    except Exception as e:
        result["errors"].append(f"PDF error: {e}")
    return result

# ═══════════════════════════════════════════════════════════════
# 6. IMAGE PROCESSOR  (PowerShell Windows OCR → Tesseract → easyocr)
# ═══════════════════════════════════════════════════════════════

# PowerShell script that calls Windows 10/11 built-in OCR engine.
# Uses WinRT via .NET reflection — works without any pip packages.
_PS_OCR_SCRIPT = r"""
Add-Type -AssemblyName System.Runtime.WindowsRuntime
$null=[Windows.Storage.StorageFile,Windows.Storage,ContentType=WindowsRuntime]
$null=[Windows.Storage.FileAccessMode,Windows.Storage,ContentType=WindowsRuntime]
$null=[Windows.Graphics.Imaging.BitmapDecoder,Windows.Graphics.Imaging,ContentType=WindowsRuntime]
$null=[Windows.Graphics.Imaging.SoftwareBitmap,Windows.Graphics.Imaging,ContentType=WindowsRuntime]
$null=[Windows.Media.Ocr.OcrEngine,Windows.Media.Ocr,ContentType=WindowsRuntime]
$null=[Windows.Storage.Streams.IRandomAccessStream,Windows.Storage.Streams,ContentType=WindowsRuntime]
$am=([System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {
    $_.Name -eq 'AsTask' -and $_.IsGenericMethod -and
    ($_.GetGenericArguments()).Count -eq 1 -and ($_.GetParameters()).Count -eq 1
}) | Select-Object -First 1
function Wa([object]$op,[Type]$t){$tk=$am.MakeGenericMethod($t).Invoke($null,@($op));$tk.Wait(-1)|Out-Null;$tk.Result}
"""

def _powershell_ocr(filepath):
    """Use Windows built-in OCR via PowerShell — no pip packages or binary installs needed."""
    abs_path = str(Path(filepath).resolve())
    script = _PS_OCR_SCRIPT + f'\n$p="{abs_path}"\n' + r"""
$file=Wa([Windows.Storage.StorageFile]::GetFileFromPathAsync($p))([Windows.Storage.StorageFile])
$st  =Wa($file.OpenAsync([Windows.Storage.FileAccessMode]::Read))([Windows.Storage.Streams.IRandomAccessStream])
$dec =Wa([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($st))([Windows.Graphics.Imaging.BitmapDecoder])
$bmp =Wa($dec.GetSoftwareBitmapAsync())([Windows.Graphics.Imaging.SoftwareBitmap])
$eng =[Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
if($null -eq $eng){throw "Windows OCR engine not available. Check language settings."}
$res =Wa($eng.RecognizeAsync($bmp))([Windows.Media.Ocr.OcrResult])
Write-Output $res.Text
"""
    with tempfile.NamedTemporaryFile(suffix=".ps1", mode="w", delete=False, encoding="utf-8") as tf:
        tf.write(script)
        ps1 = tf.name
    try:
        r = subprocess.run(
            ["powershell", "-NonInteractive", "-NoProfile",
             "-ExecutionPolicy", "Bypass", "-File", ps1],
            capture_output=True, text=True, timeout=30
        )
        if r.returncode == 0:
            return r.stdout.strip()
        raise RuntimeError(r.stderr.strip() or "PowerShell OCR returned no output")
    finally:
        try: os.unlink(ps1)
        except: pass

def _ocr_with_claude(filepath):
    """Use Claude Vision API to extract text from images."""
    cfg = load_chatbot_config()
    api_key = cfg.get("api_key", "").strip()
    if not api_key or not ANTHROPIC_OK:
        return ""
    try:
        import base64
        with open(filepath, "rb") as f:
            img_data = base64.standard_b64encode(f.read()).decode("utf-8")
        ext = filepath.rsplit(".", 1)[-1].lower()
        media_map = {
            "jpg":  "image/jpeg",
            "jpeg": "image/jpeg",
            "png":  "image/png",
            "gif":  "image/gif",
            "bmp":  "image/png",
            "tiff": "image/png"
        }
        media_type = media_map.get(ext, "image/png")
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": img_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": (
                            "Extract all text from this image. "
                            "Focus on ticket numbers like "
                            "RITM, INC, MACM, counts, and dates. "
                            "Return only the extracted text."
                        )
                    }
                ],
            }]
        )
        return response.content[0].text
    except Exception:
        return ""

def _ocr_text(filepath):
    """
    Extract text from image.
    Priority: Claude Vision → PowerShell (Windows only)
    """
    # 1. Try Claude Vision (works on all platforms)
    if ANTHROPIC_OK:
        result = _ocr_with_claude(filepath)
        if result and result.strip():
            return result
    # 2. Windows fallback
    if POWERSHELL_OK:
        return _powershell_ocr(filepath)
    return ""

def _extract_month_from_text(text):
    """Return 'Mon-YYYY' from OCR text, or current month if no date found."""
    # Pattern: "Jan 2026", "January 2026", "Jan'26", "Mar-25" etc.
    m = re.search(
        r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
        r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"[\s'\-./]?\s*(\d{2,4})", text, re.IGNORECASE)
    if m:
        abbr_raw = m.group(1)[:3].capitalize()
        if abbr_raw in MONTH_ABBR:
            yr = int(m.group(2))
            if yr < 100: yr += 2000
            return f"{abbr_raw}-{yr}"
    # Pattern: DD/MM/YYYY or YYYY-MM-DD
    for pat, yi, mi in [
        (r"\b(\d{4})[/\-](\d{2})[/\-]\d{2}\b", 1, 2),
        (r"\b\d{1,2}[/\-](\d{1,2})[/\-](\d{4})\b", 2, 1),
    ]:
        m2 = re.search(pat, text)
        if m2:
            try:
                yr, mo = int(m2.group(yi)), int(m2.group(mi))
                if 1 <= mo <= 12 and 2000 <= yr <= 2100:
                    return f"{MONTH_ABBR[mo-1]}-{yr}"
            except (ValueError, IndexError):
                pass
    now = datetime.now()
    return f"{MONTH_ABBR[now.month-1]}-{now.year}"

def process_image(filepath, ticket_type):
    result = dict(ticket_type=ticket_type, total=0, columns_found={},
                  by_team={}, by_assignee={}, by_month={}, by_team_assignee={},
                  records=[], errors=[])
    if not OCR_OK:
        result["errors"].append(
            "OCR not available. On Windows 10/11 this should work automatically via PowerShell. "
            "Restart the app — no extra install needed."
        )
        return result
    try:
        text = _ocr_text(filepath)
        if not text or not text.strip():
            result["errors"].append(
                "OCR returned empty text. The image may be too blurry, low-resolution, "
                "or the Windows OCR language pack may not be installed."
            )
            return result

        # ── Try to find explicit ticket IDs ─────────────────────
        nums = list(set(re.findall(r"(?:RITM|INC|MACM|TASK|REQ)\d+", text, re.IGNORECASE)))
        month_key = _extract_month_from_text(text)

        if nums:
            result["total"] = len(nums)
            result["records"] = [{"number": n} for n in nums]
            result["by_team"]  = {"Extracted via OCR": len(nums)}
            result["by_month"] = {month_key: len(nums)}
        else:
            # ── Fallback: try to extract a total count from the text ─
            count = 0
            for pat in [r"total\s*[:\-]?\s*(\d+)", r"count\s*[:\-]?\s*(\d+)",
                        r"(\d+)\s+tickets?", r"(\d+)\s+items?"]:
                cm = re.search(pat, text, re.IGNORECASE)
                if cm:
                    count = int(cm.group(1)); break
            if not count:
                all_n = [int(x) for x in re.findall(r"\b(\d+)\b", text) if 1 <= int(x) <= 9999]
                count = max(all_n) if all_n else 0

            if count:
                result["total"]    = count
                result["by_team"]  = {"Extracted via OCR": count}
                result["by_month"] = {month_key: count}
                result["errors"].append(
                    f"No ticket IDs (RITM/INC/MACM) found — used count {count} "
                    f"for month {month_key}. Verify this is correct."
                )
            else:
                result["errors"].append(
                    "OCR read the image but found no ticket numbers or counts. "
                    "For best results upload as Excel/CSV instead."
                )
    except Exception as e:
        err = str(e)
        err_l = err.lower()
        if "OCR engine not available" in err or "Windows OCR engine" in err:
            result["errors"].append(
                "Windows OCR language pack not found. "
                "Go to Settings → Time & Language → Language → your language → "
                "Optional features → add 'Optical character recognition'."
            )
        elif "10054" in err or "forcibly closed" in err_l or "connectionreset" in err_l:
            result["errors"].append(
                "Network error — please RESTART the app (Ctrl+C then python app.py). "
                "The updated version uses Windows built-in OCR with no internet needed."
            )
        elif "tesseract is not installed" in err_l or "not in your path" in err_l:
            result["errors"].append(
                "Tesseract binary not found. "
                "Windows OCR via PowerShell should work automatically — restart the app."
            )
        else:
            result["errors"].append(f"Image error: {e}")
    return result

# ═══════════════════════════════════════════════════════════════
# 7. FILE DISPATCHER
# ═══════════════════════════════════════════════════════════════
def dispatch_file(filepath, ticket_type):
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext in {"xlsx","xls","csv"}: return process_excel(filepath, ticket_type)
    if ext == "pdf":                return process_pdf(filepath, ticket_type)
    return process_image(filepath, ticket_type)

def process_team_efforts(filepath):
    """Process a Team Efforts sheet.
    Classify each row as RITM / Incident / MACM / JIRA by ticket-number pattern.
    Returns {total, jira_count, ritm_count, incident_count, macm_count, by_month, records, errors}"""
    result = {"total": 0, "jira_count": 0, "ritm_count": 0,
              "incident_count": 0, "macm_count": 0,
              "by_month": {}, "records": [], "errors": []}
    try:
        ext = filepath.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            df = pd.read_csv(filepath, dtype=str, encoding="utf-8", errors="replace")
        else:
            # Read all sheets, pick the one with the most rows (same as process_excel)
            # Use context manager so the file handle is released immediately on Windows.
            with pd.ExcelFile(filepath) as xl:
                best = None
                for sh in xl.sheet_names:
                    try:
                        tmp = pd.read_excel(xl, sheet_name=sh, dtype=str)
                        if best is None or len(tmp) > len(best):
                            best = tmp
                    except Exception:
                        pass
            if best is None:
                result["errors"].append("Could not read any sheet."); return result
            df = best
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all").reset_index(drop=True)
        # If the first row looks like a header (all non-numeric), try to promote it
        if len(df) > 0:
            first = df.iloc[0]
            if all(str(v).strip() == "" or not str(v).strip().replace(".","").isdigit()
                   for v in first):
                # Check if current headers look auto-generated (0, 1, 2 or Unnamed)
                auto_hdr = all(
                    str(c).strip().startswith("Unnamed") or str(c).strip().isdigit()
                    for c in df.columns
                )
                if auto_hdr:
                    df.columns = [str(v).strip() for v in df.iloc[0]]
                    df = df.iloc[1:].reset_index(drop=True)
        df = df.fillna("")

        # ── detect ticket-type column FIRST ─────────────────────
        # e.g. "Ticket Type(INC/RITM/MACM/PTASK/PRB/JIRA)"
        # or   "Operational-JIRA Deployments Tickets Type(...)"
        type_col = None
        for col in df.columns:
            cn = col.lower().replace("-", " ").replace("_", " ")
            if any(kw in cn for kw in ["tickets type", "ticket type", "type(inc",
                                        "type (inc", "jira deployments"]):
                type_col = col; break

        # ── detect ticket-number column (exclude type_col) ───────
        num_col = None
        for col in df.columns:
            if col == type_col:
                continue
            if any(kw in col.lower() for kw in
                   ["number","ticket","task","id","item","jira","issue","key"]):
                num_col = col; break
        if num_col is None:
            # fallback: first column that isn't the type column
            for col in df.columns:
                if col != type_col:
                    num_col = col; break

        # ── detect date column ───────────────────────────────────
        date_col = None
        for col in df.columns:
            if any(kw in col.lower() for kw in
                   ["date","opened","created","start","sprint"]):
                date_col = col; break

        def classify_fallback(n):
            """Classify by ticket number pattern when no type column present."""
            u = str(n).strip().upper()
            if re.match(r'^RITM\d+', u) or re.match(r'^REQ\d+', u): return "ritm"
            if re.match(r'^INC\d+', u):                              return "incident"
            if re.match(r'^TASK\d+', u) or re.match(r'^MACM\d+', u) \
               or re.match(r'^CHG\d+', u):                           return "macm"
            return "jira"

        by_month = {}
        for _, row in df.iterrows():
            if type_col:
                # Type column present: filter purely by type value
                ttype = str(row[type_col]).strip().upper()
                if not ttype or ttype in ("NAN", "NONE"):
                    continue
                if ttype != "JIRA":
                    continue
                cat = "jira"
                num = str(row[num_col]).strip() if num_col else ""
            else:
                # No type column: require a ticket number, classify by pattern
                num = str(row[num_col]).strip() if num_col else ""
                if not num or num in ("nan", "None", ""):
                    continue
                cat = classify_fallback(num)

            result[f"{cat}_count"] += 1
            result["total"] += 1
            # Monthly breakdown for JIRA
            if cat == "jira" and date_col:
                ds = str(row[date_col]).strip()
                if ds and ds not in ("nan", "None", ""):
                    try:
                        dates = _parse_dates(pd.Series([ds]))
                        if not dates.isna().all():
                            p = dates.dt.to_period("M").astype(str).iloc[0]
                            if p and p not in ("NaT", "nan") and "-" in p:
                                yr, mn = p.split("-")
                                abbr = f"{MONTH_ABBR[int(mn)-1]}-{yr}"
                                by_month[abbr] = by_month.get(abbr, 0) + 1
                    except Exception:
                        pass
            result["records"].append({"number": num, "category": cat})
        result["by_month"] = by_month
    except Exception as e:
        result["errors"].append(str(e))
    return result

# ═══════════════════════════════════════════════════════════════
# 8. EXCEL REPORT GENERATOR  (fills reference template OR creates new)
# ═══════════════════════════════════════════════════════════════

_TMPL_HDR_PAT = re.compile(
    r"^(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
    r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"[\s'\-.]?\s*(\d{2,4})\s*$", re.IGNORECASE)

def _get_template_months(template_path):
    """Return sorted list of month keys (e.g. ['Dec-2025','Jan-2026',...])
    found in the first 15 rows of any sheet in the reference Excel."""
    months = []
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=15):
                for cell in row:
                    v = str(cell.value or "").strip()
                    m = _TMPL_HDR_PAT.match(v)
                    if not m:
                        continue
                    raw = m.group(1)
                    for abbr, full in zip(MONTH_ABBR, MONTH_FULL):
                        if raw.lower() == abbr.lower() or full.lower().startswith(raw.lower()):
                            yr = int(m.group(2))
                            if yr < 100:
                                yr += 2000
                            key = f"{abbr}-{yr}"
                            if key not in months:
                                months.append(key)
                            break
        wb.close()
    except Exception:
        pass

    def _mk_sort(mk):
        parts = mk.split("-")
        if len(parts) == 2:
            abbr, yr = parts
            mo = MONTH_ABBR.index(abbr) if abbr in MONTH_ABBR else 0
            return int(yr) * 12 + mo
        return 0

    return sorted(months, key=_mk_sort)


def _get_session_template():
    """Return path to the session-specific uploaded template, falling back to global."""
    from flask import session as _sess
    sid = _sess.get("session_id")
    if sid:
        p = UPLOAD_FOLDER / sid / "reference_template.xlsx"
        if p.exists():
            return str(p)
    global_tmpl = REFERENCE_FOLDER / "reference.xlsx"
    if global_tmpl.exists():
        return str(global_tmpl)
    return None


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hcell(ws, r, c, val, fill="1F3864", fc="FFFFFF", bold=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill      = PatternFill("solid", fgColor=fill)
    cell.font      = Font(color=fc, bold=bold, name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell

def _dcell(ws, r, c, val=""):
    cell = ws.cell(row=r, column=c, value=val)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()
    cell.font      = Font(name="Calibri", size=10)
    return cell

def _monthly_counts(data):
    """Returns {Mon-YYYY: count} e.g. {'Sep-2025': 42, 'Feb-2026': 18}.
    Preserves year so Sep-2025 and Sep-2026 stay separate."""
    out = {}
    for key, cnt in data.get("by_month", {}).items():
        ks = str(key).strip()
        try:
            if re.match(r"\d{4}-\d{2}", ks):
                dt = datetime.strptime(ks[:7], "%Y-%m")
                k = f"{MONTH_ABBR[dt.month - 1]}-{dt.year}"
                out[k] = out.get(k, 0) + cnt; continue
        except Exception: pass
        kl = ks.lower()
        matched = False
        for abbr, full in zip(MONTH_ABBR, MONTH_FULL):
            if kl == abbr.lower() or kl == full.lower():
                out[ks] = out.get(ks, 0) + cnt; matched = True; break
        if not matched:
            out[ks] = out.get(ks, 0) + cnt
    return out

def _abbr_from_key(k):
    """Extract 3-letter abbreviation from 'Sep-2025', 'Sep', 'September', etc."""
    if not k: return ""
    parts = str(k).split("-")
    candidate = parts[0].strip()
    if candidate in MONTH_ABBR: return candidate
    for abbr, full in zip(MONTH_ABBR, MONTH_FULL):
        if candidate.lower() == full.lower(): return abbr
    return candidate

def _month_count(monthly_data, abbr, selected_month=None):
    """Count for a month column.
    selected_month='Sep-2025' → exact key lookup.
    selected_month=None       → sum all years for that abbreviation."""
    if selected_month:
        return monthly_data.get(selected_month, 0)
    return sum(v for k, v in monthly_data.items()
               if _abbr_from_key(k).lower() == abbr.lower())

def _months_for_display(data_list):
    """Return sorted unique 'Mon-YYYY' strings from a list of data dicts."""
    months = set()
    for data in data_list:
        if not data: continue
        for k in data.get("by_month", {}).keys():
            ks = str(k).strip()
            try:
                if re.match(r"\d{4}-\d{2}", ks):
                    dt = datetime.strptime(ks[:7], "%Y-%m")
                    months.add(f"{MONTH_ABBR[dt.month - 1]}-{dt.year}")
                    continue
            except Exception: pass
            months.add(ks)
    return sorted(months)

def fill_reference_excel(template_path, output_path, ritm, incident, macm, selected_months=None, macm_label=None, team_efforts=None):
    if not os.path.exists(template_path):
        return {"success": False, "message": f"Template not found: {template_path}"}
    try:
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
    except Exception as e:
        return {"success": False, "message": f"Cannot open template: {e}"}

    # ── Pre-compute counts ─────────────────────────────────────
    rm = _monthly_counts(ritm)
    im = _monthly_counts(incident)
    mm = _monthly_counts(macm)

    r_total = ritm.get("total", 0)
    i_total = incident.get("total", 0)
    m_total = macm.get("total", 0)

    # DB patching (Application Enhancements) — subset of RITMs
    db_monthly = _db_patching_monthly(ritm) if ritm else {}
    db_total   = sum(db_monthly.values())

    # JIRA from Team Efforts
    te = team_efforts or {}
    jm_raw = te.get("by_month") or {}
    # by_month from process_team_efforts uses Mon-YYYY keys already; normalise just in case
    jm = {}
    for k, v in jm_raw.items():
        ks = str(k).strip()
        try:
            if re.match(r"\d{4}-\d{2}", ks):
                dt = datetime.strptime(ks[:7], "%Y-%m")
                nk = f"{MONTH_ABBR[dt.month - 1]}-{dt.year}"
                jm[nk] = jm.get(nk, 0) + int(v)
                continue
        except Exception:
            pass
        jm[ks] = jm.get(ks, 0) + int(v)
    jira_total = te.get("jira_count", 0) or sum(jm.values())

    grand   = r_total + i_total + m_total + jira_total
    all_months = set(list(rm.keys()) + list(im.keys()) + list(mm.keys()) + list(jm.keys()) + list(db_monthly.keys()))
    grand_monthly = {mo: rm.get(mo, 0) + im.get(mo, 0) + mm.get(mo, 0) + jm.get(mo, 0) for mo in all_months}

    # MACM row: use user-selected label or fall back to both spellings
    if macm_label and macm_label.strip():
        macm_rules = [(macm_label.strip().lower(), mm, m_total)]
    else:
        macm_rules = [
            ("operational - application ehancements",  mm, m_total),  # template typo
            ("operational - application enhancements", mm, m_total),  # correct spelling
        ]

    # Each rule: (label, monthly_data, total_count, _unused)
    # When selected_months is set, ALL rows only write to selected month columns.
    # The total column is only written when no month filter is active.
    LABEL_RULES = [
        ("operational - incident management", im, i_total, True),
        ("incident management",               im, i_total, True),
        ("incident",                          im, i_total, True),
        ("operational - request management",  rm, r_total, True),
        ("request management",                rm, r_total, True),
        ("ritm",                              rm, r_total, True),
        *[(lbl, dat, tot, True) for lbl, dat, tot in macm_rules],
        ("macm",                              mm, m_total, True),
        ("operational-jira",                  jm, jira_total, True),
        ("operational - jira",                jm, jira_total, True),
        ("jira",                              jm, jira_total, True),
        ("application enhancements",          db_monthly, db_total,  False),  # always all months
        ("operational",                       grand_monthly, grand,  False),  # exact match — keep last
    ]

    # ── Month header parser: handles Dec'25, Jan'26, Feb 2026, March-25 etc. ──
    _HDR_PAT = re.compile(
        r"^(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
        r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"[\s'\-.]?\s*(\d{2,4})\s*$", re.IGNORECASE)

    def _parse_hdr(v):
        """Return (abbr, year_int) from 'Dec'25', 'Jan'26', 'February 2026', etc.
        Returns (abbr, None) for plain month names without year."""
        v = str(v or "").strip()
        m = _HDR_PAT.match(v)
        if m:
            raw = m.group(1)
            for abbr, full in zip(MONTH_ABBR, MONTH_FULL):
                if raw.lower() == abbr.lower() or full.lower().startswith(raw.lower()):
                    yr = int(m.group(2))
                    if yr < 100: yr += 2000
                    return (abbr, yr)
        for abbr, full in zip(MONTH_ABBR, MONTH_FULL):
            if v.lower() in (abbr.lower(), full.lower()):
                return (abbr, None)
        return None

    def _safe_write(ws, row, col, value, blocked_cols=()):
        """Write to a cell only if it is not a MergedCell and not a Forecast column."""
        if col in blocked_cols:
            return False
        c = ws.cell(row=row, column=col)
        if c.__class__.__name__ != "MergedCell":
            c.value = value  # type: ignore[assignment]
            return True
        return False

    filled = []
    for ws in wb.worksheets:

        # ── Build header map for first 15 rows ─────────────────
        hmap = {}   # (row, col) → string value
        for hrow in ws.iter_rows(max_row=15):
            for hcell in hrow:
                if hcell.value is not None:
                    hmap[(hcell.row, hcell.column)] = str(hcell.value).strip()

        # ── Detect month columns → {month_key: actual_col_idx} ─
        # month_key = "Dec-2025" (year known) or "Dec" (no year in header)
        # actual_col = the "Actual" sub-column, or the month column itself
        month_cols    = {}   # month_key → actual column index
        forecast_cols_map = {}   # month_key → forecast column index (NEVER written to)
        all_forecast_col_set = set()  # all col indices that are Forecast — blocked from writing

        for (r, c), v in hmap.items():
            parsed = _parse_hdr(v)
            if not parsed:
                continue
            abbr, year = parsed
            key = f"{abbr}-{year}" if year else abbr

            # Search for "Actual" and "Forecast" labels in same/adjacent rows.
            # Search ONLY to the right (c, c+1 … c+4) — never left.
            # The month header cell is the leftmost column of its merged range,
            # so Forecast/Actual sub-headers are always at or to the right.
            # Searching left (c-1) would accidentally pick up the *previous*
            # month's "Actual" label, mapping every month to the wrong column.
            actual_col   = None
            forecast_col = None
            for sr in [r, r + 1, r + 2]:
                for sc in [c, c + 1, c + 2, c + 3, c + 4]:
                    lbl = hmap.get((sr, sc), "").lower()
                    if lbl == "actual"   and actual_col   is None: actual_col   = sc
                    if lbl == "forecast" and forecast_col is None: forecast_col = sc
                if actual_col and forecast_col:
                    break

            # Fall back: if no "Actual" found but "Forecast" found, Actual is the next col
            if actual_col is None and forecast_col is not None:
                actual_col = forecast_col + 1

            # Only use month column itself as last resort (no sub-headers at all)
            month_cols[key] = actual_col if actual_col else c
            if forecast_col:
                forecast_cols_map[key] = forecast_col
                all_forecast_col_set.add(forecast_col)

        # ── Find "Total" column (skip if inside month area) ────
        total_col = None
        for hrow in ws.iter_rows(max_row=15):
            for hcell in hrow:
                if str(hcell.value or "").strip().lower() in ("total", "totals", "grand total"):
                    total_col = hcell.column
                    break
            if total_col:
                break

        if not month_cols:
            continue  # no month headers found in this sheet — skip

        n_written = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cv = str(cell.value).strip().lower()

                for rule_label, monthly_data, total_count, month_filterable in LABEL_RULES:
                    if rule_label == "operational":
                        matched = (cv == "operational")
                    else:
                        matched = (cv == rule_label)

                    if not matched:
                        continue

                    # Fill matching month columns (Actual only, Forecast untouched)
                    for month_key, actual_col in month_cols.items():
                        col_abbr = _abbr_from_key(month_key)

                        # ── Month filter: skip columns not in the selected set ──
                        # When months are selected, ONLY write those month columns
                        # regardless of which row type it is.
                        if selected_months:
                            col_matched = False
                            for sm in selected_months:
                                sel_abbr = _abbr_from_key(sm)
                                if col_abbr.lower() != sel_abbr.lower():
                                    continue
                                if "-" in month_key and "-" in sm:
                                    if month_key == sm:
                                        col_matched = True; break
                                else:
                                    col_matched = True; break
                            if not col_matched:
                                continue

                        # Exact count for this month column
                        exact_key = month_key if "-" in month_key else None
                        count = _month_count(monthly_data, col_abbr, exact_key)
                        if _safe_write(ws, cell.row, actual_col, count,
                                       blocked_cols=all_forecast_col_set):
                            n_written += 1

                    # Total column: only write when filling all months (no filter active)
                    if total_col and not selected_months:
                        if _safe_write(ws, cell.row, total_col, total_count,
                                       blocked_cols=all_forecast_col_set):
                            n_written += 1
                    break  # stop checking LABEL_RULES for this cell

        # ── Carry-forward pass: non-ticket rows → copy previous month Actual ──
        # Rows NOT matched by any LABEL_RULE (e.g. Capacity, Available, headcount)
        # get their selected month's Actual filled from the preceding month's value.
        # Ticket-type rows (Incident, RITM, MACM, JIRA, DB Patching, Operational)
        # are skipped — those are handled by the LABEL_RULES loop above.
        if selected_months and month_cols:
            # Sort by column index (sheet order) — NOT by calendar order.
            # This ensures Dec that appears before Jan in the sheet is treated
            # as the predecessor of Jan, regardless of month number.
            sorted_mkeys = sorted(month_cols.keys(), key=lambda mk: month_cols[mk])
            # Map each month key → the actual column of its immediately preceding month
            prev_actual_col = {}
            for idx, mk in enumerate(sorted_mkeys):
                if idx > 0:
                    prev_actual_col[mk] = month_cols[sorted_mkeys[idx - 1]]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cv = str(cell.value).strip().lower()
                    # Skip header rows
                    if cell.row <= 15:
                        continue
                    # Skip rows matched by LABEL_RULES (ticket rows — already filled above)
                    is_ticket_row = False
                    for rl, _, _, _ in LABEL_RULES:
                        if rl == "operational":
                            if cv == "operational":
                                is_ticket_row = True; break
                        else:
                            if cv == rl:
                                is_ticket_row = True; break
                    if is_ticket_row:
                        continue
                    # Skip very short / numeric labels unlikely to be row labels
                    if len(cv) < 3:
                        continue

                    # For each selected month, copy previous month's Actual value
                    for sm in selected_months:
                        sel_abbr = _abbr_from_key(sm).lower()
                        for mk, actual_col in month_cols.items():
                            col_abbr = _abbr_from_key(mk).lower()
                            if col_abbr != sel_abbr:
                                continue
                            if "-" in mk and "-" in sm and mk != sm:
                                continue
                            prev_col = prev_actual_col.get(mk)
                            if prev_col is None:
                                break
                            prev_val = ws.cell(row=cell.row, column=prev_col).value
                            if prev_val is not None:
                                _safe_write(ws, cell.row, actual_col,
                                            prev_val, blocked_cols=all_forecast_col_set)
                            break

        if n_written:
            filled.append(ws.title)

    if not filled:
        # No sheet had detectable month columns — create a summary sheet
        ws2 = wb.create_sheet("Ticket Summary")
        for ci, h in enumerate(["Ticket Type", "Total"] + MONTH_ABBR, 1):
            ws2.cell(row=1, column=ci, value=h).font = Font(bold=True)
        for ri, (lbl, dat, tot) in enumerate([
            ("Operational - Incident Management",    im,           i_total),
            ("Operational - Request Management",     rm,           r_total),
            ("Operational - Application Enhancements", mm,         m_total),
            ("Operational-JIRA",                     jm,           jira_total),
            ("Application Enhancements (DB Patching)", db_monthly, db_total),
            ("Operational (Total)",                  grand_monthly, grand),
        ], 2):
            ws2.cell(row=ri, column=1, value=lbl)
            ws2.cell(row=ri, column=2, value=tot)
            for ci, a in enumerate(MONTH_ABBR, 3):
                ws2.cell(row=ri, column=ci, value=_month_count(dat, a, selected_month))
        filled.append("Ticket Summary (new sheet)")

    wb.save(output_path); wb.close()
    return {"success": True, "message": f"Filled sheets: {', '.join(filled)}", "output_path": output_path}

def generate_standalone_report(output_path, ritm, incident, macm, team_efforts=None, selected_months=None):
    """Generate a standalone Excel report.

    selected_months: list like ['Mar-2026'] — when set, RITM/Incident/MACM/JIRA
    columns are filtered to only those months.  DB Patching and other derived
    rows always show ALL months so historical data is preserved.
    """
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(color="FFFFFF", bold=True)

    te         = team_efforts or {}
    jira_total = te.get("jira_count", 0)
    db_monthly = _db_patching_monthly(ritm) if ritm else {}
    db_total   = sum(db_monthly.values())

    # ── Helper: total respecting selected_months filter ───────
    def _total_sel(data):
        """Total count for selected months only (or full total when no filter)."""
        if not data:
            return 0
        if not selected_months:
            return data.get("total", 0)
        bm = _monthly_counts(data)
        return sum(bm.get(sm, 0) for sm in selected_months)

    # ── Normalise JIRA monthly data ────────────────────────────
    jm_raw = te.get("by_month") or {}
    jm_norm = {}
    for k, v in jm_raw.items():
        ks = str(k).strip()
        try:
            if re.match(r"\d{4}-\d{2}", ks):
                dt = datetime.strptime(ks[:7], "%Y-%m")
                nk = f"{MONTH_ABBR[dt.month-1]}-{dt.year}"
                jm_norm[nk] = jm_norm.get(nk, 0) + int(v); continue
        except Exception: pass
        jm_norm[ks] = jm_norm.get(ks, 0) + int(v)

    # Effective per-type totals (respect selected_months for RITM/INC/MACM/JIRA)
    r_eff = _total_sel(ritm)
    i_eff = _total_sel(incident)
    m_eff = _total_sel(macm)
    if selected_months and te:
        j_eff = sum(jm_norm.get(sm, 0) for sm in selected_months)
    else:
        j_eff = jira_total

    # ── Summary sheet ──────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "ServiceNow Ticket Count Report"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    month_label = (", ".join(selected_months) if selected_months else "All Months")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Period: {month_label}"
    for ci, h in enumerate(["Work Type","Total"],1):
        c = ws.cell(row=4,column=ci,value=h); c.font=hfont; c.fill=hfill
    rows = [
        ("Operational - Request Management",       r_eff),
        ("Operational - Incident Management",       i_eff),
        ("Operational - Application Enhancements",  m_eff),
        ("Operational-JIRA",                        j_eff),
        ("Application Enhancements (DB Patching)",  db_total),  # always all months
    ]
    for ri, (lbl, tot) in enumerate(rows, 5):
        ws.cell(row=ri,column=1,value=lbl); ws.cell(row=ri,column=2,value=tot)
    grand_row = 5 + len(rows)
    grand_total = r_eff + i_eff + m_eff + j_eff
    ws.cell(row=grand_row,column=1,value="GRAND TOTAL").font=Font(bold=True)
    ws.cell(row=grand_row,column=2,value=grand_total).font=Font(bold=True)
    ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=15

    # ── Helper: get month-filtered team/assignee counts ────────
    def _filtered_by_key(data, key_type):
        """Return {name: count} filtered by selected_months if active."""
        if not selected_months:
            return data.get(key_type, {})
        bm_key = "by_month_team" if key_type == "by_team" else "by_month_assignee"
        bm = data.get(bm_key, {})
        counts = {}
        for sm in selected_months:
            # by_month_team keys are "YYYY-MM"; selected_months are "Mon-YYYY"
            parts = sm.split("-")
            if len(parts) == 2:
                try:
                    mo = MONTH_ABBR.index(parts[0]) + 1
                    iso_key = f"{parts[1]}-{mo:02d}"
                except (ValueError, IndexError):
                    iso_key = sm
            else:
                iso_key = sm
            for name, cnt in bm.get(iso_key, {}).items():
                counts[name] = counts.get(name, 0) + cnt
        return counts

    # ── By Team ────────────────────────────────────────────────
    wt = wb.create_sheet("By Team")
    for ci,h in enumerate(["Team","RITM","Incident","MACM","Total"],1):
        c=wt.cell(row=1,column=ci,value=h); c.font=hfont; c.fill=hfill
    r_team = _filtered_by_key(ritm, "by_team")
    i_team = _filtered_by_key(incident, "by_team")
    m_team = _filtered_by_key(macm, "by_team")
    all_teams = sorted(set(list(r_team.keys())+list(i_team.keys())+list(m_team.keys())))
    for ri,t in enumerate(all_teams,2):
        r=r_team.get(t,0); i=i_team.get(t,0); m=m_team.get(t,0)
        for ci,v in enumerate([t,r,i,m,r+i+m],1): wt.cell(row=ri,column=ci,value=v)
    for col in "ABCDE": wt.column_dimensions[col].width=25

    # ── By Assignee ────────────────────────────────────────────
    wa = wb.create_sheet("By Assignee")
    for ci,h in enumerate(["Assignee","RITM","Incident","MACM","Total"],1):
        c=wa.cell(row=1,column=ci,value=h); c.font=hfont; c.fill=hfill
    r_asgn = _filtered_by_key(ritm, "by_assignee")
    i_asgn = _filtered_by_key(incident, "by_assignee")
    m_asgn = _filtered_by_key(macm, "by_assignee")
    all_asgn = sorted(set(list(r_asgn.keys())+list(i_asgn.keys())+list(m_asgn.keys())))
    for ri,a in enumerate(all_asgn,2):
        r=r_asgn.get(a,0); i=i_asgn.get(a,0); m=m_asgn.get(a,0)
        for ci,v in enumerate([a,r,i,m,r+i+m],1): wa.cell(row=ri,column=ci,value=v)
    for col in "ABCDE": wa.column_dimensions[col].width=25

    # ── Monthly sheet ──────────────────────────────────────────
    # RITM/Incident/MACM/JIRA: only show selected months when a filter is active.
    # DB Patching: always all months (preserve historical data).
    rm = _monthly_counts(ritm)
    im = _monthly_counts(incident)
    mm = _monthly_counts(macm)

    # All months from all sources (used for DB Patching row)
    all_months_full = sorted(set(
        list(rm.keys()) + list(im.keys()) + list(mm.keys()) +
        list(jm_norm.keys()) + list(db_monthly.keys())
    ))

    # Months to show for RITM/INC/MACM/JIRA rows
    ticket_months = selected_months if selected_months else all_months_full

    # Union set for the Monthly sheet rows (ticket months + db months)
    monthly_display = sorted(set(ticket_months) | set(db_monthly.keys()))

    wm = wb.create_sheet("Monthly")
    for ci,h in enumerate(["Month","RITM","Incident","MACM","JIRA","DB Patching","Total"],1):
        c=wm.cell(row=1,column=ci,value=h); c.font=hfont; c.fill=hfill
    for ri, mo in enumerate(monthly_display, 2):
        r = rm.get(mo, 0) if mo in (ticket_months or []) else 0
        i = im.get(mo, 0) if mo in (ticket_months or []) else 0
        m = mm.get(mo, 0) if mo in (ticket_months or []) else 0
        j = jm_norm.get(mo, 0) if mo in (ticket_months or []) else 0
        d = db_monthly.get(mo, 0)   # always full data
        for ci, v in enumerate([mo, r, i, m, j, d, r+i+m+j], 1):
            wm.cell(row=ri, column=ci, value=v)
    for col in ["A","B","C","D","E","F","G"]: wm.column_dimensions[col].width=15

    wb.save(output_path); wb.close()
    return output_path

# ═══════════════════════════════════════════════════════════════
# 9. REFERENCE TEMPLATE CREATOR
# ═══════════════════════════════════════════════════════════════
def create_reference_template(path):
    if os.path.exists(path): return
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for sheet_name, label in [
        ("RITM",     "Operational - Request Management"),
        ("Incident", "Operational - Incident Management"),
        ("MACM",     "Operational - Application Enhancements"),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:Z1")
        ws["A1"].value = f"ServiceNow {sheet_name} — Monthly Count"
        ws["A1"].font  = Font(bold=True, size=13, color="1F3864")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28
        _hcell(ws, 2, 1, "Ticket Type", "2E75B6")
        col = 2
        for mo in MONTH_ABBR:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
            _hcell(ws, 2, col, mo, "1F3864")
            _hcell(ws, 3, col,   "Actual",   "2E75B6")
            _hcell(ws, 3, col+1, "Forecast", "375623")
            col += 2
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        _hcell(ws, 2, col, "Total", "7030A0")
        _hcell(ws, 3, col, "Actual", "7030A0"); _hcell(ws, 3, col+1, "Forecast", "7030A0")
        _dcell(ws, 4, 1, label)
        c2 = 2
        for _ in MONTH_ABBR:
            _dcell(ws, 4, c2, 0); _dcell(ws, 4, c2+1, 0); c2 += 2
        _dcell(ws, 4, c2, 0); _dcell(ws, 4, c2+1, 0)
        ws.column_dimensions["A"].width = 16
        for i in range(2, c2+2): ws.column_dimensions[get_column_letter(i)].width = 9

    ws_s = wb.create_sheet("Summary")
    ws_s.sheet_view.showGridLines = False
    ws_s.merge_cells("A1:P1")
    ws_s["A1"].value = "Monthly Ticket Summary"
    ws_s["A1"].font  = Font(bold=True, size=13, color="1F3864")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    _hcell(ws_s, 2, 1, "Ticket Type", "2E75B6")
    for ci, mo in enumerate(MONTH_ABBR, 2): _hcell(ws_s, 2, ci, mo, "1F3864")
    _hcell(ws_s, 2, 14, "Total", "7030A0")
    for ri, lbl in enumerate(["Operational - Request Management","Operational - Incident Management","Operational - Application Enhancements"], 3):
        _dcell(ws_s, ri, 1, lbl)
        for ci in range(2, 15): _dcell(ws_s, ri, ci, 0)
    ws_s.column_dimensions["A"].width = 16
    for i in range(2, 15): ws_s.column_dimensions[get_column_letter(i)].width = 10

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path); wb.close()
    print(f"Reference template created: {path}")

# ═══════════════════════════════════════════════════════════════
# 10. SESSION HELPERS
# ═══════════════════════════════════════════════════════════════
def _ensure_session():
    if not session.get("session_id"):
        session["session_id"] = str(uuid.uuid4())
    return session["session_id"]

def _save(key, data):
    sid = _ensure_session()
    d = UPLOAD_FOLDER / sid; d.mkdir(exist_ok=True)
    with open(d / f"{key}.json", "w") as f: json.dump(data, f)

def _load(key):
    sid = session.get("session_id")
    if not sid: return {}
    p = UPLOAD_FOLDER / sid / f"{key}.json"
    if p.exists():
        with open(p) as f: return json.load(f)
    return {}

def _cleanup():
    sid = session.get("session_id")
    if sid:
        d = UPLOAD_FOLDER / sid
        if d.exists(): shutil.rmtree(d, ignore_errors=True)

# ═══════════════════════════════════════════════════════════════
# 10b. EMAIL HELPERS
# ═══════════════════════════════════════════════════════════════
def load_email_config():
    if EMAIL_CONFIG_FILE.exists():
        try:
            with open(EMAIL_CONFIG_FILE) as f:
                cfg = json.load(f)
            # Merge with defaults so new keys are always present
            merged = dict(DEFAULT_EMAIL_CONFIG); merged.update(cfg)
            return merged
        except Exception:
            pass
    return dict(DEFAULT_EMAIL_CONFIG)

def save_email_config(cfg):
    with open(EMAIL_CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# In-memory store for background email job results {job_id: result_dict}
_email_jobs: dict = {}

def _name_from_email(email):
    """Derive a display name from an email address local part.
    e.g. aileni.anilkumar@cognizant.com -> 'Aileni Anilkumar'"""
    local = email.split("@")[0]
    parts = re.split(r"[._]", local)
    return " ".join(p.capitalize() for p in parts if p)

def send_email_report(report_path, selected_month=None):
    cfg = load_email_config()
    recipients = [r.strip() for r in cfg.get("recipients", []) if r.strip()]
    if not recipients:
        return {"success": False, "error": "No recipients configured. Open Email Settings and add recipient emails."}

    use_auth = bool(cfg.get("smtp_user", "").strip() and cfg.get("smtp_password", "").strip())

    if not use_auth and not cfg.get("no_auth"):
        return {"success": False, "error":
                "No credentials set. Either enter SMTP User + Password, "
                "or enable 'No Auth (Internal Relay)' in Email Settings."}

    sender_addr = cfg.get("smtp_user", "").strip() or cfg.get("sender_email", "noreply@company.com")
    month_name  = datetime.now().strftime("%B")
    subject     = "30-60-90 Report"

    try:
        with open(report_path, "rb") as fp:
            attachment_data = fp.read()
        filename = os.path.basename(report_path)

        context = ssl.create_default_context()
        with smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"]), timeout=15) as server:
            if cfg.get("use_tls", False) and use_auth:
                server.starttls(context=context)
            if use_auth:
                server.login(cfg["smtp_user"].strip(), cfg["smtp_password"].strip())

            for recipient in recipients:
                name = _name_from_email(recipient)
                body = f"Hi {name},\n\n30-60-90 Report\n@{month_name}"

                msg = MIMEMultipart()
                msg["From"]    = f"{cfg['sender_name']} <{sender_addr}>"
                msg["To"]      = recipient
                msg["Subject"] = subject
                msg.attach(MIMEText(body, "plain"))

                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment_data)
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
                msg.attach(part)

                server.sendmail(sender_addr, [recipient], msg.as_string())

        return {"success": True, "message": f"Report sent to {', '.join(recipients)}"}
    except TimeoutError:
        return {"success": False, "error":
                f"Timed out connecting to {cfg['smtp_host']}:{cfg['smtp_port']}. "
                "Your network is blocking outgoing SMTP."}
    except smtplib.SMTPAuthenticationError as e:
        return {"success": False, "error": f"Login failed — wrong email or password. Detail: {e}"}
    except smtplib.SMTPServerDisconnected:
        return {"success": False, "error":
                f"Server {cfg['smtp_host']} closed the connection immediately."}
    except smtplib.SMTPException as e:
        return {"success": False, "error": f"SMTP error: {e}"}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ═══════════════════════════════════════════════════════════════
# 10c. CHATBOT HELPERS
# ═══════════════════════════════════════════════════════════════
def _search_ticket(number, all_data):
    """Search ritm/incident/macm records for a ticket number (exact then partial)."""
    num_clean = number.strip().upper()
    matches = []
    for ttype, data in all_data.items():
        if not data:
            continue
        for rec in data.get("records", []):
            rec_num = str(rec.get("number", "")).strip().upper()
            if rec_num == num_clean or num_clean in rec_num:
                matches.append({"ticket_type": ttype.upper(), **rec})
    return matches

def _search_by_keyword(keyword, all_data, limit=5):
    """Full-text search across short_description in all records."""
    kw = keyword.lower()
    results = []
    for ttype, data in all_data.items():
        if not data:
            continue
        for rec in data.get("records", []):
            desc = str(rec.get("short_description", "")).lower()
            if kw in desc and desc:
                results.append({"ticket_type": ttype.upper(), **rec})
    return results[:limit]

def _build_data_context(all_data):
    """Build a concise context string about loaded ticket data for the LLM."""
    lines = ["=== Loaded ServiceNow Data ==="]
    for ttype, data in all_data.items():
        if not data:
            continue
        total = data.get("total", 0)
        lines.append(f"\n{ttype.upper()}: {total} tickets")
        teams = data.get("by_team", {})
        if teams:
            top = sorted(teams.items(), key=lambda x: x[1], reverse=True)[:5]
            lines.append(f"  Top teams: {', '.join(f'{t}({c})' for t, c in top)}")
        months = data.get("by_month", {})
        if months:
            lines.append(f"  Months: {', '.join(list(months.keys())[:6])}")
    return "\n".join(lines)

def _find_duplicates(all_data, month=None):
    """Duplicate = same short_description across RITM, Incident, MACM.
    RITM & Incident use 'opened' for month filter; MACM uses 'closed'.
    Optional month='Jan-2024' filters records to that month before grouping.
    Returns list of {description, count, ritm_count, incident_count, macm_count, tickets} sorted by count desc."""
    from collections import defaultdict

    def _norm_date(val):
        s = str(val).strip()
        if s in ("", "nan", "None", "N/A", "NaT"):
            return ""
        m = re.match(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', s)
        if m:
            return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
        m = re.match(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', s)
        if m:
            y = m.group(3)
            y = "20" + y if len(y) == 2 else y
            return f"{y}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        return s[:10] if len(s) >= 10 else s

    # Convert "Jan-2024" → "2024-01" prefix for date-based month filter
    month_prefix = None
    if month and month != "all":
        parts = month.strip().split("-")
        if len(parts) == 2 and parts[0] in MONTH_ABBR and parts[1].isdigit():
            month_prefix = f"{parts[1]}-{str(MONTH_ABBR.index(parts[0]) + 1).zfill(2)}"

    # RITM & Incident → filter by opened; MACM → filter by closed
    date_field = {"ritm": "opened", "incident": "opened", "macm": "closed"}

    groups = defaultdict(list)
    for ttype in ("ritm", "incident", "macm"):
        data = all_data.get(ttype) or {}
        field = date_field[ttype]
        for rec in (data.get("records") or []):
            desc = str(rec.get("short_description", "")).strip()
            if not desc or desc in ("N/A", "nan", "None", ""):
                continue
            # Month filter: only include records whose relevant date falls in selected month
            if month_prefix:
                date = _norm_date(rec.get(field, ""))
                if not date or not date.startswith(month_prefix):
                    continue
            groups[desc.lower()].append({
                "number":      rec.get("number", ""),
                "ticket_type": ttype.upper(),
                "team":        rec.get("team", ""),
                "assignee":    rec.get("assignee", ""),
                "state":       rec.get("state", ""),
                "description": desc,
                "opened":      rec.get("opened", ""),
                "closed":      rec.get("closed", ""),
            })

    def _to_mon_year(date_str):
        """Parse any date string → 'Mon-YYYY', e.g. 'Jan-2025'. Returns '' on failure."""
        s = str(date_str).strip()
        if s in ("", "nan", "None", "N/A", "NaT"):
            return ""
        m = re.match(r'(\d{4})[.\-/](\d{1,2})', s)
        if m:
            yr, mo_num = int(m.group(1)), int(m.group(2))
            if 1 <= mo_num <= 12:
                return f"{MONTH_ABBR[mo_num - 1]}-{yr}"
        return ""

    result = []
    for _, recs in groups.items():
        if len(recs) >= 2:
            # Monthly breakdown for this group
            group_months = {}
            for r in recs:
                field = "closed" if r["ticket_type"] == "MACM" else "opened"
                mo = _to_mon_year(r.get(field, ""))
                if mo:
                    group_months[mo] = group_months.get(mo, 0) + 1
            result.append({
                "description":    recs[0]["description"],
                "count":          len(recs),
                "ritm_count":     sum(1 for r in recs if r["ticket_type"] == "RITM"),
                "incident_count": sum(1 for r in recs if r["ticket_type"] == "INCIDENT"),
                "macm_count":     sum(1 for r in recs if r["ticket_type"] == "MACM"),
                "by_month":       group_months,
                "tickets":        recs,
            })
    result.sort(key=lambda x: x["count"], reverse=True)

    # Overall monthly summary: total duplicate tickets per month
    monthly_summary = {}
    for g in result:
        for mo, cnt in g["by_month"].items():
            monthly_summary[mo] = monthly_summary.get(mo, 0) + cnt

    return result, monthly_summary

def _count_db_patching(ritm_data, month_prefix=None):
    """Count RITM records whose short_description contains 'database patch'.
    month_prefix: 'YYYY-MM' string; if set, only count records from that month."""
    count = 0
    for rec in (ritm_data.get("records") or []):
        if "database patch" not in str(rec.get("short_description", "")).lower():
            continue
        if month_prefix:
            opened = str(rec.get("opened", "")).strip()
            m = re.match(r'(\d{4})[.\-/](\d{1,2})', opened)
            if not m:
                continue
            if f"{m.group(1)}-{m.group(2).zfill(2)}" != month_prefix:
                continue
        count += 1
    return count

def _db_patching_monthly(ritm_data):
    """Return {Mon-YYYY: count} for RITM records with 'database patch' in short_description."""
    monthly = {}
    for rec in (ritm_data.get("records") or []):
        if "database patch" not in str(rec.get("short_description", "")).lower():
            continue
        opened = str(rec.get("opened", "")).strip()
        m = re.match(r'(\d{4})[.\-/](\d{1,2})', opened)
        if not m:
            continue
        yr, mo_num = int(m.group(1)), int(m.group(2))
        if 1 <= mo_num <= 12:
            key = f"{MONTH_ABBR[mo_num - 1]}-{yr}"
            monthly[key] = monthly.get(key, 0) + 1
    return monthly

def _get_top10_slow(ritm_data, incident_data):
    """Combine RITM and Incident top10_slow lists and return overall top-10 by duration."""
    combined = []
    for src in [ritm_data, incident_data]:
        if not src:
            continue
        for rec in (src.get("top10_slow") or []):
            combined.append(rec)
    combined.sort(key=lambda x: x.get("duration_days", 0), reverse=True)
    return combined[:10]

def _extract_subject(desc):
    """Extract the meaningful subject (system/app/tool) from a ticket description."""
    if not desc:
        return "the reported issue"
    stop = {
        "please","need","needs","require","requires","request","requests","the","a","an",
        "for","to","in","on","with","of","by","at","is","are","was","has","have","been",
        "install","installed","installing","setup","configure","reset","access","unable",
        "enable","disable","create","delete","remove","update","change","get","getting",
        "user","users","laptop","computer","pc","device","system","my","our","their",
        "new","old","current","issue","problem","error","not","working","broken","failed",
        "help","please","asap","urgent","support","ticket","request","hi","hello",
    }
    words = re.findall(r"[A-Za-z0-9][A-Za-z0-9\.\-\_\/]*", desc)
    subject_words = [w for w in words if w.lower() not in stop and len(w) > 2]
    if subject_words:
        return " ".join(subject_words[:4])
    return desc[:50].rstrip(" .,")

def _dynamic_resolution(ticket):
    """Generate contextual, step-by-step resolution using the actual ticket fields."""
    desc     = str(ticket.get("short_description", "")).strip()
    state    = str(ticket.get("state",    "")).strip()
    priority = str(ticket.get("priority", "")).strip()
    team     = str(ticket.get("team",     "")).strip()
    dl       = desc.lower()
    subject  = _extract_subject(desc)
    lines    = []

    # ── Status / priority banner ─────────────────────────────────
    closed_states = {"closed", "resolved", "cancelled", "complete", "completed", "cancel"}
    if state.lower() in closed_states:
        lines.append(f"**Status:** {state} — steps shown for reference / repeat-issue handling.")
    elif priority:
        p = priority.lower()
        if any(x in p for x in ["1 -", "p1", "critical"]):
            lines.append(f"**Priority:** {priority} — escalate to **{team or 'owning team'}** immediately.")
        elif any(x in p for x in ["2 -", "p2", "high"]):
            lines.append(f"**Priority:** {priority} — begin within the hour.")

    # ── Category detection → specific numbered steps ─────────────
    if any(w in dl for w in ["install", "deploy", "deployment", "software install", "app install"]):
        lines += [
            f"**Resolution — Software Deployment ({subject})**",
            f"1. Confirm **{subject}** is listed in the approved software catalogue",
            f"2. Check available licence count for **{subject}** in the asset portal",
            f"3. Push **{subject}** to the target device via SCCM / Intune silent install",
            f"4. Log into the device post-install: confirm **{subject}** launches and shows correct version",
            f"5. Update the CMDB software asset record for **{subject}** (device, user, date)",
            f"6. Notify the requester with confirmation and close the ticket",
        ]
    elif any(w in dl for w in ["password", "pwd", "forgot password", "locked out", "lockout", "lock out", "unlock"]):
        lines += [
            f"**Resolution — Password / Account Unlock ({subject})**",
            f"1. Verify identity: confirm employee ID + manager name or HR record",
            f"2. Check account status in AD: `Get-ADUser <username> -Properties LockedOut,PasswordExpired`",
            f"3. Unlock if locked: `Unlock-ADAccount -Identity <username>`",
            f"4. Reset password: `Set-ADAccountPassword -Identity <username> -NewPassword (Read-Host -AsSecureString)`",
            f"5. Force change at next logon: `Set-ADUser <username> -ChangePasswordAtLogon $true`",
            f"6. Re-register MFA if the user's device was also replaced; confirm sign-in, then close",
        ]
    elif any(w in dl for w in ["access", "permission", "role", "privilege", "entitlement", "group", "membership"]):
        lines += [
            f"**Resolution — Access / Permissions ({subject})**",
            f"1. Validate written approval from line manager + data/system owner for **{subject}**",
            f"2. Cross-check the entitlement matrix: confirm correct AD group or application role for **{subject}**",
            f"3. Add user to the AD group or provision the role in **{subject}**",
            f"4. Ask the user to verify access: confirm they can reach **{subject}** as expected",
            f"5. Update the CMDB access register: record grant date, approver, and system (**{subject}**)",
            f"6. Send confirmation email to the requester and close the ticket",
        ]
    elif any(w in dl for w in ["vpn", "remote access", "remote desktop", "rdp", "citrix", "zscaler"]):
        lines += [
            f"**Resolution — Remote Access ({subject})**",
            f"1. Confirm **{subject}** client version: compare against the latest release on the vendor portal",
            f"2. Check user certificate validity: `certmgr.msc` → Personal → re-enrol if expired",
            f"3. Test on a different network (mobile hotspot) to isolate home ISP / firewall issues",
            f"4. Flush **{subject}** credential cache: remove saved profiles and re-authenticate",
            f"5. Review firewall ACL for the user's source IP range — confirm split-tunnel rules for **{subject}**",
            f"6. Escalate to Network team with packet capture and client logs if unresolved",
        ]
    elif any(w in dl for w in ["email", "outlook", "mailbox", "calendar", "exchange", "mail flow", "o365", "m365", "teams"]):
        lines += [
            f"**Resolution — Email / Collaboration ({subject})**",
            f"1. Check mailbox quota: `Get-MailboxStatistics <email> | fl TotalItemSize,ItemCount`",
            f"2. Run Message Trace in Exchange Admin Centre to confirm mail flow for affected messages",
            f"3. Re-create Outlook profile: Control Panel → Mail → Show Profiles → Add new profile",
            f"4. Check junk / spam rules: add sender to Safe Senders; verify transport rules in EAC",
            f"5. Verify Autodiscover DNS and EWS endpoint: `Test-OutlookWebServices -Identity <email>`",
            f"6. Escalate to Exchange/M365 team with EWS trace and message ID if flow issue persists",
        ]
    elif any(w in dl for w in ["slow", "performance", "freeze", "crash", "not responding", "hang", "restart", "reboot", "bsod", "blue screen"]):
        lines += [
            f"**Resolution — Performance / Stability ({subject})**",
            f"1. Check disk space: `Get-PSDrive C | Select Used,Free` — archive data if <15% free",
            f"2. Task Manager → Performance: check CPU and RAM peaks; kill top offending processes",
            f"3. Run full AV scan: confirm definitions are current and scan completes clean",
            f"4. Review Event Viewer: `eventvwr.msc` → Windows Logs → filter Critical/Error (last 24h)",
            f"5. Run system file check: `sfc /scannow` then `DISM /Online /Cleanup-Image /RestoreHealth`",
            f"6. If hardware fault suspected (thermal, RAM, disk): escalate to hardware team with diagnostic logs",
        ]
    elif any(w in dl for w in ["network", "internet", "connectivity", "wifi", "wireless", "lan", "ethernet", "no internet", "cannot connect"]):
        lines += [
            f"**Resolution — Network Connectivity ({subject})**",
            f"1. Ping default gateway: `ping <gateway>` — confirms Layer 3 reachability from the device",
            f"2. Release/renew DHCP: `ipconfig /release` then `ipconfig /renew`",
            f"3. Flush DNS cache: `ipconfig /flushdns` — retest connectivity after flush",
            f"4. Check NIC driver version: Device Manager → update if >6 months old",
            f"5. Inspect switch port: confirm correct VLAN assignment and check for interface errors in switch logs",
            f"6. Escalate to Network team with MAC address, switch hostname, port ID, and ping results",
        ]
    elif any(w in dl for w in ["printer", "print", "printing", "scanner", "scan to email", "scan to folder"]):
        lines += [
            f"**Resolution — Print / Scan ({subject})**",
            f"1. Clear the print queue: `services.msc` → Print Spooler → Stop → delete all files in `C:\\Windows\\System32\\spool\\PRINTERS` → Start",
            f"2. Remove and re-add the printer: Settings → Printers & Scanners → Add a device",
            f"3. Re-install driver: download exact model driver from vendor site (avoid generic PCL6 if possible)",
            f"4. Test print from a different device to isolate user laptop vs. printer fault",
            f"5. Check toner level, paper jam, and any error code on the printer's front panel",
            f"6. Log hardware fault ticket with Facilities if the issue is physical; otherwise close after confirmed fix",
        ]
    elif any(w in dl for w in ["sap", "erp", "hana", "fiori", "sap gui", "bapi", "idoc"]):
        lines += [
            f"**Resolution — SAP / ERP ({subject})**",
            f"1. Check SAP system status: transaction **SM51** (Application Servers) — confirm all nodes are green",
            f"2. Verify user roles: **SU01** → Roles tab — check role name, validity dates, and profile status",
            f"3. Clear SAP GUI cache: navigate to `%APPDATA%\\SAP\\Common` and delete cache files",
            f"4. Review background job queue: **SM37** — identify any failed jobs impacting the user's workload",
            f"5. Check system log: **SM21** (last 30 min) — collect dump and send to SAP Basis team",
            f"6. Raise SAP OSS note / incident if confirmed as a SAP product defect",
        ]
    elif any(w in dl for w in ["mfa", "multi-factor", "authenticator", "2fa", "two factor", "azure ad", "okta", "sso", "single sign"]):
        lines += [
            f"**Resolution — MFA / SSO ({subject})**",
            f"1. Disable MFA temporarily in Azure AD / Okta for the user (document reason and approver)",
            f"2. User re-registers: open Microsoft Authenticator → Add account → Scan QR from the MFA portal",
            f"3. If device is lost: issue a one-time bypass code (valid 10 min) from the admin portal",
            f"4. Review Conditional Access policies: check location/device compliance requirements for **{subject}**",
            f"5. Re-enable MFA and perform a sign-in test with the new registration",
            f"6. Document re-registration timestamp and approver in the ticket for audit trail; close",
        ]
    elif any(w in dl for w in ["new user", "onboard", "joiner", "new account", "create user", "new starter", "provision user", "new employee"]):
        lines += [
            f"**Resolution — User Onboarding ({subject})**",
            f"1. Confirm signed HR approval + confirmed start date from the requesting manager",
            f"2. Create AD account: set correct OU, UPN (`firstname.lastname@domain.com`), display name, job title, manager",
            f"3. Add to standard department groups (email DL, Teams channel, shared drives, VPN group)",
            f"4. Assign M365 licence via Azure AD group-based licensing; confirm mailbox provisioned",
            f"5. Enrol MFA: send self-service registration link to the user's personal email",
            f"6. Prepare device (image, standard apps, department-specific software) and deliver with welcome pack",
        ]
    else:
        lines += [
            f"**Resolution — {subject}**",
            f"1. Contact the requester to confirm exact symptoms and reproduce the issue for **{subject}**",
            f"2. Collect evidence: error message screenshot, affected user/device, time of occurrence",
            f"3. Search the knowledge base and open incidents for known issues with **{subject}**",
            f"4. Apply the standard runbook for **{subject}** (check KB / Confluence / vendor docs)",
            f"5. If unresolved after runbook: escalate to the owning team for **{subject}** with full diagnostics",
            f"6. Document every action taken in the ticket; close only after requester confirms resolution",
        ]
    return "\n".join(lines)

def _answer_stats(all_data, message):
    """Answer statistics questions dynamically from real loaded ticket data."""
    r = all_data.get("ritm", {}) or {}
    i = all_data.get("incident", {}) or {}
    m = all_data.get("macm", {}) or {}
    r_total = r.get("total", 0)
    i_total = i.get("total", 0)
    m_total = m.get("total", 0)
    grand   = r_total + i_total + m_total

    if not grand:
        return "No ticket data loaded yet. Upload your RITM, Incident, and MACM files first."

    msg = message.lower()
    parts = []

    if any(w in msg for w in ["team", "group", "assignment group", "who handles", "busiest team"]):
        all_teams = {}
        for src in [r, i, m]:
            for team, cnt in (src.get("by_team", {}) or {}).items():
                all_teams[team] = all_teams.get(team, 0) + cnt
        top = sorted(all_teams.items(), key=lambda x: x[1], reverse=True)[:10]
        if top:
            max_cnt = top[0][1] or 1
            parts.append("**Top Assignment Groups (all ticket types combined):**\n")
            for rank, (team, cnt) in enumerate(top, 1):
                bar = "\u2588" * max(1, round(cnt / max_cnt * 12))
                pct = round(cnt / grand * 100)
                parts.append(f"{rank}. **{team}**\n   {cnt} tickets ({pct}% of total)  `{bar}`")

    elif any(w in msg for w in ["assignee", "person", "who", "agent", "technician", "engineer", "worked by", "resolved by"]):
        all_people = {}
        for src in [r, i, m]:
            for person, cnt in (src.get("by_assignee", {}) or {}).items():
                all_people[person] = all_people.get(person, 0) + cnt
        top = sorted(all_people.items(), key=lambda x: x[1], reverse=True)[:10]
        if top:
            parts.append("**Top Assignees (all ticket types):**\n")
            for rank, (person, cnt) in enumerate(top, 1):
                r_cnt = (r.get("by_assignee") or {}).get(person, 0)
                i_cnt = (i.get("by_assignee") or {}).get(person, 0)
                m_cnt = (m.get("by_assignee") or {}).get(person, 0)
                parts.append(f"{rank}. **{person}** — {cnt} tickets  (RITM: {r_cnt} | Inc: {i_cnt} | MACM: {m_cnt})")

    elif any(w in msg for w in ["month", "trend", "monthly", "jan","feb","mar","apr","may","jun",
                                  "jul","aug","sep","oct","nov","dec","this month","last month"]):
        rm = _monthly_counts(r); im = _monthly_counts(i); mm = _monthly_counts(m)
        all_months_set = set(list(rm.keys()) + list(im.keys()) + list(mm.keys()))
        def _sort_key(k):
            parts_k = k.split("-")
            try:
                yr = int(parts_k[1]) if len(parts_k) > 1 else 0
                mo = MONTH_ABBR.index(parts_k[0]) if parts_k[0] in MONTH_ABBR else 0
            except Exception:
                yr, mo = 0, 0
            return (yr, mo)
        months_sorted = sorted(all_months_set, key=_sort_key)
        parts.append("**Monthly Ticket Breakdown:**\n")
        for mo in months_sorted:
            r_cnt = rm.get(mo, 0); i_cnt = im.get(mo, 0); m_cnt = mm.get(mo, 0)
            tot   = r_cnt + i_cnt + m_cnt
            bar   = "\u2588" * max(1, round(tot / max((rm.get(months_sorted[-1],0)+im.get(months_sorted[-1],0)+mm.get(months_sorted[-1],0)), tot, 1) * 10))
            parts.append(f"**{mo}:** {tot} total  `{bar}`\n   RITM: {r_cnt} | Incidents: {i_cnt} | MACM: {m_cnt}")

    else:
        # General summary with real numbers
        parts.append("**Ticket Summary — Loaded Data**\n")
        parts.append(f"• RITM (Request Items): **{r_total}**")
        parts.append(f"• Incidents: **{i_total}**")
        parts.append(f"• MACM (App Enhancements): **{m_total}**")
        parts.append(f"• **Grand Total: {grand}**")

        # Top team overall
        all_teams = {}
        for src in [r, i, m]:
            for team, cnt in (src.get("by_team", {}) or {}).items():
                all_teams[team] = all_teams.get(team, 0) + cnt
        if all_teams:
            top_team, top_cnt = max(all_teams.items(), key=lambda x: x[1])
            parts.append(f"\nTop Team: **{top_team}** — {top_cnt} tickets")

        # Latest month
        rm = _monthly_counts(r); im = _monthly_counts(i); mm = _monthly_counts(m)
        all_mo = set(list(rm.keys()) + list(im.keys()) + list(mm.keys()))
        if all_mo:
            latest = sorted(all_mo)[-1]
            l_tot  = rm.get(latest, 0) + im.get(latest, 0) + mm.get(latest, 0)
            parts.append(f"Latest Month: **{latest}** — {l_tot} tickets")

        parts.append("\nAsk me: `top teams`, `monthly trend`, `top assignees`")

    return "\n".join(parts)

def _answer_top10_slow(ritm_data, incident_data):
    """Return a formatted chatbot answer listing top-10 slowest tickets with 3 resolution methods each."""
    top10 = _get_top10_slow(ritm_data or {}, incident_data or {})
    if not top10:
        return ("No duration data found. Make sure your RITM and Incident files "
                "contain both an **opened/start date** column and a **closed/resolved date** column.")

    lines = [f"**Top {len(top10)} Slowest Tickets — by Resolution Time**\n"]
    for rank, rec in enumerate(top10, 1):
        num   = rec.get("number", "?")
        ttype = rec.get("ticket_type", "?")
        desc  = rec.get("short_description", "No description")
        team  = rec.get("team", "")
        asgn  = rec.get("assignee", "")
        dur   = rec.get("duration_days", "?")
        opened = rec.get("opened", "")
        closed = rec.get("closed", "")

        lines.append(f"---\n**#{rank} — {num}** ({ttype})  ·  **{dur} days**")
        lines.append(f"**Desc:** {desc[:100]}")
        if team:  lines.append(f"**Team:** {team}")
        if asgn:  lines.append(f"**Assignee:** {asgn}")
        if opened: lines.append(f"**Opened:** {opened}  →  **Closed:** {closed}")

        # Generate 3 resolution methods based on PRIORITY
        prio_raw = rec.get("priority", "") or ""
        prio_l   = prio_raw.lower().strip()
        subject  = _extract_subject(desc)

        if any(p in prio_l for p in ["1", "critical", "p1"]):
            methods = [
                f"1. 🚨 **Immediate War-Room Bridge**: Open a P1 bridge call NOW with all resolver groups for **{num}**. Assign a dedicated incident commander; every 15 min update to stakeholders until resolved.",
                f"2. 🔴 **Executive Escalation**: Notify IT leadership and business owners within 30 minutes. If no resolution progress in 1 hour, escalate to vendor/third-party SLA emergency clause.",
                f"3. 📋 **Post-Incident RCA**: Within 48 hrs of closure, hold a blameless post-mortem for **{num}**. Raise a Problem ticket, document the KEDB entry, and implement a permanent fix to prevent recurrence.",
            ]
        elif any(p in prio_l for p in ["2", "high", "p2"]):
            methods = [
                f"1. ⚠️ **Senior Engineer Assignment**: Assign a senior/L3 engineer to **{num}** within 1 hour. Team lead must be notified immediately; daily progress check-in mandatory.",
                f"2. 🔔 **4-Hour SLA Checkpoint**: If not resolved in 4 hours, auto-escalate **{num}** to the next resolver tier and cc the service delivery manager for visibility.",
                f"3. 🛠️ **Targeted RCA**: Perform root-cause analysis within 24 hrs of resolution. Update the knowledge base with the fix procedure so L2 can handle similar issues independently next time.",
            ]
        elif any(p in prio_l for p in ["3", "medium", "p3", "moderate"]):
            methods = [
                f"1. 📅 **Scheduled Resolution Slot**: Assign **{num}** to the appropriate team queue with a firm commitment date. Use the standard change/request process; no emergency change needed.",
                f"2. 📊 **SLA Breach-Risk Report**: Include **{num}** in the weekly SLA breacher report. If it approaches the breach threshold, bump priority and reassign to an available engineer.",
                f"3. 📖 **Knowledge Article**: After resolution, create or update a KB article for the issue type in **{num}** so L1 agents can self-serve similar cases and reduce future ticket volume.",
            ]
        else:
            # P4 / Low / unknown
            methods = [
                f"1. 🟢 **Self-Service / KB Deflection**: Check if there is an existing KB article for **{num}**'s issue type. Route the user to the self-service portal to resolve independently, reducing queue load.",
                f"2. 📥 **L1 Bulk Processing**: Handle **{num}** in the next L1 processing batch (daily/weekly). Group similar low-priority tickets together for efficient bulk resolution.",
                f"3. ✅ **Auto-Close Candidate**: If the ticket is stale (no user response in 5+ business days), trigger an auto-close workflow with a closure notification and KB link for future reference.",
            ]

        lines.append("**3 Resolution Methods:**")
        lines.extend(methods)

    lines.append("\n*Tip: Ask me for a specific ticket number (e.g. RITM0001234) for detailed resolution steps.*")
    return "\n\n".join(lines)

def _generate_chat_response(message, tickets_found, data_context, all_data, history):
    """Try Claude API (haiku) → dynamic data-driven fallback."""
    cfg     = load_chatbot_config()
    api_key = cfg.get("api_key", "").strip()

    # Build rich ticket context for Claude
    ticket_ctx = ""
    if tickets_found:
        rows = []
        for t in tickets_found[:3]:
            rows.append(
                f"Ticket: {t.get('number','?')} ({t.get('ticket_type','?')})\n"
                f"Description: {t.get('short_description','N/A')}\n"
                f"Team: {t.get('team','N/A')} | Assignee: {t.get('assignee','N/A')}\n"
                f"State: {t.get('state','N/A')} | Priority: {t.get('priority','N/A')}\n"
                f"Opened: {t.get('opened','N/A')} | Closed: {t.get('closed','N/A')}"
            )
        ticket_ctx = "\n\n".join(rows)

    if api_key and ANTHROPIC_OK:
        try:
            client = anthropic.Anthropic(api_key=api_key)
            # Include top-10 slowest context for AI awareness
            top10_ctx = ""
            top10 = _get_top10_slow(
                all_data.get("ritm", {}), all_data.get("incident", {}))
            if top10:
                t10_lines = ["Top 10 Slowest Tickets (by resolution time):"]
                for idx, rec in enumerate(top10, 1):
                    t10_lines.append(
                        f"{idx}. {rec.get('number','?')} ({rec.get('ticket_type','?')}) "
                        f"— {rec.get('duration_days','?')} days — {rec.get('short_description','')[:60]}"
                    )
                top10_ctx = "\n".join(t10_lines)

            system = (
                "You are a ServiceNow IT support expert embedded in a ticket reporting tool. "
                "When given a ticket, show its key details and provide numbered, actionable resolution steps "
                "that reference the specific system or application mentioned in the description. "
                "For stats questions, calculate and answer from the data context. "
                "When asked about slow/long/top-10 tickets, provide 3 practical resolution methods per ticket. "
                "Be concise and professional. Use **bold** for emphasis and numbered steps for resolutions."
                f"\n\nLoaded ticket data:\n{data_context}"
            )
            if top10_ctx:
                system += f"\n\n{top10_ctx}"
            if ticket_ctx:
                system += f"\n\nMatched ticket(s):\n{ticket_ctx}"

            msgs = []
            for h in history[-6:]:
                if h.get("role") in ("user", "assistant"):
                    msgs.append({"role": h["role"], "content": h["content"]})
            if msgs and msgs[-1]["role"] == "user" and msgs[-1]["content"] == message:
                msgs = msgs[:-1]
            msgs.append({"role": "user", "content": message})

            response = client.messages.create(
                model=cfg.get("model", "claude-haiku-4-5-20251001"),
                max_tokens=700,
                system=system,
                messages=msgs
            )
            return response.content[0].text
        except Exception:
            pass  # fall through to dynamic fallback

    # ── Dynamic data-driven fallback (no API key) ────────────────
    if tickets_found:
        t = tickets_found[0]
        parts = []

        # ── Ticket header ──────────────────────────────────────────
        num   = t.get("number", "?")
        ttype = t.get("ticket_type", "?")
        parts.append(f"**{num}** ({ttype})")

        # ── All available fields from the actual record ────────────
        _state_val = str(t.get("state", "")).lower()
        _resolver_label = "Resolved By" if any(w in _state_val for w in ("closed", "resolved", "complete")) else "Assignee"
        field_order = [
            ("Description",   "short_description"),
            ("State",         "state"),
            ("Priority",      "priority"),
            ("Team",          "team"),
            (_resolver_label, "assignee"),
            ("Opened",        "opened"),
            ("Closed",        "closed"),
        ]
        detail_lines = []
        for label, key in field_order:
            val = str(t.get(key, "")).strip()
            if val and val not in ("N/A", "nan", "None", ""):
                detail_lines.append(f"**{label}:** {val}")
        if detail_lines:
            parts.append("\n".join(detail_lines))

        # ── Note if multiple matches ───────────────────────────────
        if len(tickets_found) > 1:
            others = [tx.get("number", "?") for tx in tickets_found[1:4]]
            parts.append(f"*Also matched: {', '.join(others)}*")

        # ── Dynamic resolution based on actual description ─────────
        parts.append(_dynamic_resolution(t))

        parts.append("*Add your Anthropic API key in Settings (⚙) for AI-powered contextual analysis.*")
        return "\n\n".join(parts)

    # ── Stats / summary questions ────────────────────────────────
    msg_lower = message.lower()
    stats_words = ["how many", "total", "count", "summary", "stats", "overview",
                   "top team", "top assignee", "which team", "most tickets",
                   "busiest", "monthly", "month", "trend", "who handled"]
    if any(w in msg_lower for w in stats_words):
        return _answer_stats(all_data, message)

    # ── No data loaded ───────────────────────────────────────────
    has_data = any(d.get("total") for d in all_data.values() if d)
    if not has_data:
        return ("No ticket data loaded yet. Upload your RITM, Incident, and MACM files first.\n\n"
                "Once loaded I can:\n"
                "• Look up any ticket by number: `RITM0001234` or `INC0045678`\n"
                "• Search by keyword: `password reset`, `VPN`, `SAP`\n"
                "• Answer stats: `top teams`, `monthly trend`, `how many incidents`\n"
                "• Generate step-by-step resolutions from the actual ticket description")

    # ── No match found ───────────────────────────────────────────
    # Try to suggest similar tickets from keyword
    suggestions = _search_by_keyword(message, all_data, limit=3)
    if suggestions:
        lines = [f"No exact match for **{message}**. Similar tickets found:\n"]
        for s in suggestions:
            lines.append(
                f"• **{s.get('number','?')}** ({s.get('ticket_type','?')}) — "
                f"{str(s.get('short_description',''))[:80]}"
            )
        lines.append("\nType the ticket number to get its full details and resolution.")
        return "\n".join(lines)

    return (f"No ticket matching **{message}** found in the loaded data.\n\n"
            "Try:\n"
            "• Full ticket number: `RITM0001234` or `INC0045678`\n"
            "• Keyword from the description: `VPN`, `access request`, `Adobe`\n"
            "• Stats question: `top teams`, `how many RITM this month?`")

# ═══════════════════════════════════════════════════════════════
# 11. CHATBOT WIDGET  (injected into all pages)
# ═══════════════════════════════════════════════════════════════
CHATBOT_WIDGET = """
<!-- ── Ticket Assistant Chatbot ─────────────────────────────── -->
<style>
@keyframes cb-bounce{0%,80%,100%{transform:translateY(0)}40%{transform:translateY(-6px)}}
@keyframes cb-fadein{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
@keyframes cb-blink{50%{opacity:0}}

#cb-launcher{position:fixed;bottom:24px;left:24px;z-index:9999;width:56px;height:56px;
  border-radius:50%;background:linear-gradient(135deg,#0d6efd,#6f42c1);
  box-shadow:0 4px 20px rgba(13,110,253,.45);display:flex;align-items:center;
  justify-content:center;cursor:pointer;color:#fff;font-size:1.5rem;
  transition:transform .2s,box-shadow .2s;border:none;outline:none;}
#cb-launcher:hover{transform:scale(1.1);box-shadow:0 6px 28px rgba(13,110,253,.6);}
#cb-launcher .cb-badge{position:absolute;top:-3px;right:-3px;background:#dc3545;
  color:#fff;font-size:.6rem;font-weight:700;border-radius:50%;width:18px;height:18px;
  display:none;align-items:center;justify-content:center;border:2px solid #fff;}

#cb-panel{position:fixed;bottom:92px;left:24px;z-index:9999;width:390px;
  display:none;flex-direction:column;border-radius:18px;
  box-shadow:0 10px 48px rgba(0,0,0,.24);background:#fff;overflow:hidden;
  font-family:'Segoe UI',system-ui,sans-serif;max-height:560px;
  animation:cb-fadein .22s ease;}
.cb-header{background:linear-gradient(135deg,#0d6efd,#6f42c1);padding:14px 16px;
  display:flex;align-items:center;gap:10px;color:#fff;flex-shrink:0;}
.cb-avatar-bot{width:36px;height:36px;border-radius:50%;background:rgba(255,255,255,.22);
  display:flex;align-items:center;justify-content:center;font-size:1.1rem;flex-shrink:0;}
.cb-hbtn{background:transparent;border:none;color:#fff;font-size:1rem;cursor:pointer;
  padding:4px 8px;border-radius:6px;transition:background .15s;}
.cb-hbtn:hover{background:rgba(255,255,255,.22);}
.cb-online-dot{width:8px;height:8px;border-radius:50%;background:#4ade80;
  display:inline-block;margin-right:4px;box-shadow:0 0 0 2px rgba(74,222,128,.3);}

#cb-settings-pane{display:none;padding:11px 14px;background:#f8f9fa;
  border-bottom:1px solid #dee2e6;flex-shrink:0;}

#cb-messages{flex:1;overflow-y:auto;padding:14px 12px;display:flex;
  flex-direction:column;gap:4px;min-height:180px;max-height:340px;}
#cb-messages::-webkit-scrollbar{width:4px}
#cb-messages::-webkit-scrollbar-thumb{background:#dee2e6;border-radius:4px}

/* message rows */
.cb-row{display:flex;align-items:flex-end;gap:7px;animation:cb-fadein .18s ease;}
.cb-row.cb-row-user{flex-direction:row-reverse;}
.cb-row-av{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:.85rem;flex-shrink:0;}
.cb-row-av.bot{background:linear-gradient(135deg,#0d6efd22,#6f42c122);color:#0d6efd;}
.cb-row-av.user{background:linear-gradient(135deg,#0d6efd,#6f42c1);color:#fff;}
.cb-msg-wrap{display:flex;flex-direction:column;max-width:85%;}
.cb-row.cb-row-user .cb-msg-wrap{align-items:flex-end;}
.cb-user-msg{background:linear-gradient(135deg,#0d6efd,#6f42c1);
  color:#fff;padding:9px 13px;border-radius:16px 16px 4px 16px;
  font-size:.84rem;line-height:1.48;word-break:break-word;}
.cb-bot-msg{background:#f0f4ff;color:#1e293b;
  padding:9px 13px;border-radius:16px 16px 16px 4px;
  font-size:.84rem;line-height:1.56;word-break:break-word;
  border-left:3px solid #0d6efd;}
.cb-ts{font-size:.65rem;color:#adb5bd;margin-top:2px;padding:0 2px;}

/* quick chips */
#cb-chips{padding:6px 12px 2px;display:flex;flex-wrap:wrap;gap:6px;flex-shrink:0;
  border-top:1px solid #f0f0f0;}
.cb-chip{padding:5px 11px;border:1.5px solid #c7d2fe;border-radius:50px;
  font-size:.75rem;color:#4338ca;cursor:pointer;white-space:nowrap;
  background:#eef2ff;transition:background .15s,border-color .15s;font-family:inherit;}
.cb-chip:hover{background:#c7d2fe;border-color:#818cf8;color:#1e1b4b;}

/* typing */
#cb-typing{display:none;padding:2px 14px 8px;font-size:.8rem;color:#6c757d;flex-shrink:0;}
.cb-dot{width:7px;height:7px;border-radius:50%;background:#adb5bd;display:inline-block;
  animation:cb-bounce .9s infinite ease-in-out;}

/* input */
#cb-input-row{padding:9px 11px;border-top:1px solid #f0f0f0;
  display:flex;gap:8px;align-items:flex-end;flex-shrink:0;}
#cb-input{flex:1;border:1.5px solid #dee2e6;border-radius:10px;padding:8px 11px;
  font-size:.84rem;resize:none;outline:none;font-family:inherit;
  max-height:80px;overflow-y:auto;line-height:1.4;}
#cb-input:focus{border-color:#0d6efd;}
#cb-send{padding:8px 13px;background:linear-gradient(135deg,#0d6efd,#6f42c1);
  color:#fff;border:none;border-radius:10px;font-size:1rem;cursor:pointer;
  transition:opacity .2s,transform .15s;flex-shrink:0;}
#cb-send:hover{opacity:.88;transform:scale(1.05);}

/* maximize */
#cb-panel.cb-maximized{width:min(92vw,820px);max-height:min(90vh,720px);}
#cb-panel.cb-maximized #cb-messages{max-height:calc(min(90vh,720px) - 230px);}

/* cursor blink for typewriter */
.cb-cursor{display:inline-block;width:2px;height:.85em;background:#0d6efd;
  margin-left:1px;vertical-align:text-bottom;animation:cb-blink .7s step-end infinite;}
</style>

<button id="cb-launcher" onclick="cbToggle()" title="Ticket Assistant">
  🤖<span class="cb-badge" id="cb-notif-badge">1</span>
</button>

<div id="cb-panel">
  <div class="cb-header">
    <div class="cb-avatar-bot">🤖</div>
    <div style="flex:1">
      <div style="font-weight:700;font-size:.94rem">Ticket Assistant</div>
      <div style="font-size:.72rem;opacity:.88">
        <span class="cb-online-dot"></span>Online · AI-powered resolutions
      </div>
    </div>
    <button class="cb-hbtn" onclick="cbOpenSettings()" title="API Key">⚙️</button>
    <button class="cb-hbtn" id="cb-maxbtn" onclick="cbMaximize()" title="Maximize">⤢</button>
    <button class="cb-hbtn" onclick="cbToggle()" title="Close">✕</button>
  </div>

  <div id="cb-settings-pane">
    <div style="font-size:.78rem;font-weight:700;color:#495057;margin-bottom:7px">
      ⚙️ Claude API Key — enables AI-powered resolutions
    </div>
    <div style="display:flex;gap:7px;align-items:center">
      <input id="cb-api-key" type="password" placeholder="sk-ant-..."
        style="flex:1;padding:6px 10px;border:1px solid #ced4da;border-radius:8px;
               font-size:.8rem;outline:none;font-family:inherit"/>
      <button onclick="cbSaveSettings()"
        style="padding:6px 13px;background:#0d6efd;color:#fff;border:none;
               border-radius:8px;font-size:.8rem;cursor:pointer;white-space:nowrap">Save</button>
    </div>
    <div id="cb-settings-status" style="font-size:.73rem;margin-top:5px;color:#198754;min-height:16px"></div>
    <div style="font-size:.7rem;color:#6c757d;margin-top:2px">
      Get your key at console.anthropic.com · Uses claude-haiku (fast &amp; cost-efficient)
    </div>
  </div>

  <div id="cb-messages"></div>

  <div id="cb-typing">
    <div style="display:flex;align-items:center;gap:6px;padding-left:35px">
      <div class="cb-row-av bot" style="width:24px;height:24px;font-size:.75rem">🤖</div>
      <span style="display:inline-flex;gap:4px;align-items:center;
        background:#f0f4ff;padding:7px 12px;border-radius:12px;border-left:3px solid #0d6efd">
        <span class="cb-dot"></span>
        <span class="cb-dot" style="animation-delay:.2s"></span>
        <span class="cb-dot" style="animation-delay:.4s"></span>
      </span>
    </div>
  </div>

  <div id="cb-chips">
    <button class="cb-chip" onclick="cbChip('Top 10 slowest tickets')">⏱ Top 10 Slowest</button>
    <button class="cb-chip" onclick="cbChip('Monthly summary')">📅 Monthly Summary</button>
    <button class="cb-chip" onclick="cbChip('Top teams by volume')">👥 Top Teams</button>
    <button class="cb-chip" onclick="cbChip('Top assignees')">🏆 Top Assignees</button>
  </div>

  <div id="cb-input-row">
    <textarea id="cb-input" rows="1"
      placeholder="Ask me anything about your tickets…"
      onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();cbSend();}"
      oninput="this.style.height='auto';this.style.height=this.scrollHeight+'px'"></textarea>
    <button id="cb-send" onclick="cbSend()" title="Send">➤</button>
  </div>
</div>

<script>
(function(){
  const CB_HIST=[];
  let _cbOpen=false, _cbMax=false, _cbGreeted=false;

  /* ── helpers ── */
  function _ts(){
    const n=new Date();
    return n.getHours().toString().padStart(2,'0')+':'+n.getMinutes().toString().padStart(2,'0');
  }

  function _cbFmt(t){
    return String(t)
      .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
      .replace(/\\*\\*(.*?)\\*\\*/g,'<strong>$1</strong>')
      .replace(/`([^`]+)`/g,'<code style="background:#e8ecf0;padding:1px 5px;border-radius:4px;font-size:.8rem;font-family:monospace">$1</code>')
      .replace(/^(\\u2022|•) (.+)$/gm,'<div style="padding-left:12px">• $2</div>')
      .replace(/\\n/g,'<br>');
  }

  function _addRow(role, html){
    const box=document.getElementById('cb-messages');
    const row=document.createElement('div');
    row.className='cb-row'+(role==='user'?' cb-row-user':'');

    const av=document.createElement('div');
    av.className='cb-row-av '+(role==='user'?'user':'bot');
    av.textContent=role==='user'?'👤':'🤖';

    const wrap=document.createElement('div');
    wrap.className='cb-msg-wrap';

    const bubble=document.createElement('div');
    bubble.className=role==='user'?'cb-user-msg':'cb-bot-msg';
    bubble.innerHTML=html;

    const ts=document.createElement('div');
    ts.className='cb-ts';
    ts.textContent=_ts();

    wrap.appendChild(bubble);
    wrap.appendChild(ts);
    row.appendChild(av);
    row.appendChild(wrap);
    box.appendChild(row);
    box.scrollTop=box.scrollHeight;
    return bubble;
  }

  /* ── typewriter effect ── */
  function _typewrite(bubble, fullHtml, onDone){
    // Strip HTML tags for char-by-char reveal; then swap to formatted at end
    const plain=fullHtml.replace(/<[^>]+>/g,'').replace(/&amp;/g,'&').replace(/&lt;/g,'<').replace(/&gt;/g,'>');
    let i=0;
    const cursor=document.createElement('span');
    cursor.className='cb-cursor';
    bubble.innerHTML='';
    bubble.appendChild(cursor);
    const box=document.getElementById('cb-messages');
    const speed=plain.length>400?6:plain.length>150?10:14; // adaptive speed
    function tick(){
      if(i<plain.length){
        cursor.insertAdjacentText('beforebegin',plain[i]);
        i++;
        box.scrollTop=box.scrollHeight;
        setTimeout(tick,speed);
      } else {
        // swap in fully formatted html
        bubble.innerHTML=fullHtml;
        box.scrollTop=box.scrollHeight;
        if(onDone) onDone();
      }
    }
    tick();
  }

  function _cbAddBot(text, animate){
    const html=_cbFmt(text);
    const bubble=_addRow('bot',animate?'':html);
    if(animate) _typewrite(bubble, html);
    CB_HIST.push({role:'assistant',content:text});
  }

  function _cbAddUser(text){
    _addRow('user',_cbFmt(text));
    CB_HIST.push({role:'user',content:text});
  }

  /* ── chips ── */
  window.cbChip=function(txt){
    document.getElementById('cb-input').value=txt;
    cbSend();
  };

  /* ── maximize ── */
  window.cbMaximize=function(){
    _cbMax=!_cbMax;
    const p=document.getElementById('cb-panel');
    const btn=document.getElementById('cb-maxbtn');
    p.classList.toggle('cb-maximized',_cbMax);
    btn.textContent=_cbMax?'\u2921':'\u2922';
    btn.title=_cbMax?'Restore':'Maximize';
    const msgs=document.getElementById('cb-messages');
    if(msgs) msgs.scrollTop=msgs.scrollHeight;
  };

  /* ── toggle ── */
  window.cbToggle=function(){
    _cbOpen=!_cbOpen;
    const p=document.getElementById('cb-panel');
    p.style.display=_cbOpen?'flex':'none';
    // hide unread badge
    const badge=document.getElementById('cb-notif-badge');
    if(badge) badge.style.display='none';
    if(_cbOpen){
      if(!_cbGreeted){
        _cbGreeted=true;
        setTimeout(()=>{
          _cbAddBot(
            "Hey there! 👋 I\u2019m your **Ticket Assistant**.\\n\\n"+
            "Here\u2019s what I can help you with:\\n"+
            "\u2022 **Top 10 slowest tickets** \u2014 with priority-based resolution steps\\n"+
            "\u2022 **Lookup** any ticket by number (e.g. `RITM0001234`)\\n"+
            "\u2022 **Search** by keyword (e.g. `VPN access`, `password reset`)\\n"+
            "\u2022 **Stats** \u2014 monthly summary, top teams, top assignees\\n\\n"+
            "Try one of the quick options below, or just type your question!",
            true
          );
        }, 350);
      }
      setTimeout(()=>document.getElementById('cb-input').focus(),80);
    }
  };

  /* ── send ── */
  window.cbSend=async function(){
    const inp=document.getElementById('cb-input');
    const msg=inp.value.trim(); if(!msg) return;
    inp.value=''; inp.style.height='auto';
    _cbAddUser(msg);
    const typ=document.getElementById('cb-typing');
    typ.style.display='block';
    document.getElementById('cb-messages').scrollTop=99999;
    try{
      const r=await fetch('/api/chat',{
        method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({message:msg,history:CB_HIST.slice(-8)})
      });
      const d=await r.json();
      typ.style.display='none';
      _cbAddBot(d.reply||'Sorry, I could not process that.', true);
    }catch(e){
      typ.style.display='none';
      _cbAddBot('Oops \u2014 connection error. Please check your network and try again.', false);
    }
  };

  /* ── settings ── */
  window.cbOpenSettings=function(){
    const s=document.getElementById('cb-settings-pane');
    const open=s.style.display!=='none';
    s.style.display=open?'none':'block';
    if(!open){
      fetch('/api/chatbot-config').then(r=>r.json()).then(d=>{
        const k=document.getElementById('cb-api-key');
        if(k) k.placeholder=d.has_key?'API key saved \u2713 (enter new to replace)':'sk-ant-...';
        const st=document.getElementById('cb-settings-status');
        if(st) st.textContent=d.has_key?'\u2713 API key configured \u2014 AI responses active':'No API key \u2014 using rule-based fallback';
        if(st) st.style.color=d.has_key?'#198754':'#fd7e14';
      }).catch(()=>{});
    }
  };

  window.cbSaveSettings=async function(){
    const key=document.getElementById('cb-api-key').value.trim();
    const st=document.getElementById('cb-settings-status');
    if(!key){st.textContent='Please enter your API key.';st.style.color='#dc3545';return;}
    st.textContent='Saving\u2026'; st.style.color='#6c757d';
    try{
      const r=await fetch('/api/chatbot-config',{
        method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({api_key:key})
      });
      const d=await r.json();
      if(d.success){
        st.textContent='\u2713 Saved! AI responses now active.';
        st.style.color='#198754';
        document.getElementById('cb-api-key').value='';
        document.getElementById('cb-api-key').placeholder='API key saved \u2713';
      } else {
        st.textContent='\u2717 Save failed'; st.style.color='#dc3545';
      }
    }catch(e){st.textContent='\u2717 Error: '+e.message; st.style.color='#dc3545';}
  };

  /* show badge on load to hint the chatbot exists */
  window.addEventListener('load',()=>{
    const b=document.getElementById('cb-notif-badge');
    if(b){b.style.display='flex'; setTimeout(()=>b.style.display='none',8000);}
  });
})();
</script>
<!-- ── End Chatbot Widget ──────────────────────────────────── -->
"""

# ═══════════════════════════════════════════════════════════════
# 12. CSS  (loaded from style.css)
# ═══════════════════════════════════════════════════════════════
CSS = (Path(__file__).parent / "style.css").read_text(encoding="utf-8")

# ═══════════════════════════════════════════════════════════════
# 13. APP JS  (loaded from app.js)
# ═══════════════════════════════════════════════════════════════
APP_JS = (Path(__file__).parent / "app.js").read_text(encoding="utf-8")

# ═══════════════════════════════════════════════════════════════
# 14. DASHBOARD JS  (loaded from dashboard.js)
# ═══════════════════════════════════════════════════════════════
DASHBOARD_JS = (Path(__file__).parent / "dashboard.js").read_text(encoding="utf-8")

# ═══════════════════════════════════════════════════════════════
# 14. HTML TEMPLATES
# ═══════════════════════════════════════════════════════════════
INDEX_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>ServiceNow Ticket Counter</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet"/>
<link rel="stylesheet" href="/assets/style.css"/>
</head>
<body class="bg-light">
<nav class="navbar navbar-dark" style="background:linear-gradient(135deg,#1a1a2e,#16213e,#0f3460)">
  <div class="container-fluid">
    <span class="navbar-brand fw-bold fs-4"><i class="bi bi-ticket-perforated me-2 text-info"></i>ServiceNow Ticket Counter</span>
    <a href="/reset" class="btn btn-outline-light btn-sm"><i class="bi bi-arrow-counterclockwise me-1"></i>Reset</a>
  </div>
</nav>

<div class="container py-4">

<!-- Progress Steps -->
<div class="card shadow-sm mb-4 border-0">
  <div class="card-body p-4">
    <h6 class="text-muted text-uppercase fw-bold mb-3 small">Upload Progress</h6>
    <div class="d-flex align-items-center">
      {% for n,label in [(1,"RITM"),(2,"Incident"),(3,"MACM"),(4,"Team Efforts"),(5,"Reference"),(6,"Results")] %}
        <div class="flex-fill text-center">
          <div class="step-circle {% if step > n %}done{% elif step < n %}disabled{% endif %}" id="step{{n}}-circle">
            <span class="snum">{{n}}</span><i class="bi bi-check-lg scheck"></i>
          </div>
          <div class="mt-2 fw-semibold small {% if step > n %}text-success{% elif step == n %}text-primary{% else %}text-muted{% endif %}" id="step{{n}}-label">{{label}}</div>
          <div class="text-muted small" id="step{{n}}-count">{% if step > n %}Done{% elif step == n %}Active{% else %}Waiting{% endif %}</div>
        </div>
        {% if n < 6 %}<div class="step-connector flex-fill {% if step > n %}active{% endif %}" id="conn{{n}}"></div>{% endif %}
      {% endfor %}
    </div>
  </div>
</div>

<!-- RITM Upload -->
<div id="upload-ritm" class="upload-section {% if step != 1 %}d-none{% endif %}">
  <div class="card shadow border-0 upload-card ritm-card">
    <div class="card-body p-4">
      <div class="d-flex align-items-center mb-3">
        <div class="icon-badge bg-primary text-white rounded-3 p-3 me-3"><i class="bi bi-file-earmark-spreadsheet fs-3"></i></div>
        <div><h4 class="mb-0 fw-bold">Step 1 — Upload RITM File</h4><p class="text-muted mb-0 small">ServiceNow RITM (Request Item) export</p></div>
      </div>
      <div class="drop-zone" id="drop-ritm" onclick="document.getElementById('file-ritm').click()">
        <i class="bi bi-cloud-upload fs-1 text-primary mb-2"></i>
        <p class="fw-semibold mb-1">Drag &amp; drop your RITM file here</p>
        <p class="text-muted small mb-0">or click to browse &nbsp;·&nbsp; Excel · CSV · PDF · Images</p>
        <input type="file" id="file-ritm" class="d-none" accept=".xlsx,.xls,.csv,.pdf,.png,.jpg,.jpeg,.bmp,.tiff"/>
      </div>
      <div id="ritm-file-info" class="mt-3 d-none">
        <i class="bi bi-file-check text-success me-2"></i>
        <span id="ritm-filename" class="fw-semibold"></span>
        <span class="badge bg-success ms-2" id="ritm-size"></span>
      </div>
      <div id="ritm-progress" class="mt-3 d-none">
        <div class="progress" style="height:8px"><div class="progress-bar progress-bar-striped progress-bar-animated bg-primary" style="width:100%"></div></div>
        <p class="text-muted small mt-1">Processing…</p>
      </div>
      <div id="ritm-result" class="mt-3 d-none"></div>
      <div class="mt-3">
        <button class="btn btn-primary btn-lg px-5" id="btn-upload-ritm" disabled onclick="uploadFile('ritm')">
          <i class="bi bi-upload me-2"></i>Upload &amp; Process RITM
        </button>
      </div>
    </div>
  </div>
</div>

<!-- Incident Upload -->
<div id="upload-incident" class="upload-section {% if step != 2 %}d-none{% endif %}">
  <div class="card shadow border-0 upload-card incident-card">
    <div class="card-body p-4">
      <div class="d-flex align-items-center mb-3">
        <div class="icon-badge bg-warning text-dark rounded-3 p-3 me-3"><i class="bi bi-exclamation-triangle fs-3"></i></div>
        <div><h4 class="mb-0 fw-bold">Step 2 — Upload Incident File</h4><p class="text-muted mb-0 small">ServiceNow Incident export</p></div>
      </div>
      <div class="drop-zone" id="drop-incident" onclick="document.getElementById('file-incident').click()">
        <i class="bi bi-cloud-upload fs-1 text-warning mb-2"></i>
        <p class="fw-semibold mb-1">Drag &amp; drop your Incident file here</p>
        <p class="text-muted small mb-0">or click to browse &nbsp;·&nbsp; Excel · CSV · PDF · Images</p>
        <input type="file" id="file-incident" class="d-none" accept=".xlsx,.xls,.csv,.pdf,.png,.jpg,.jpeg,.bmp,.tiff"/>
      </div>
      <div id="incident-file-info" class="mt-3 d-none">
        <i class="bi bi-file-check text-success me-2"></i>
        <span id="incident-filename" class="fw-semibold"></span>
        <span class="badge bg-success ms-2" id="incident-size"></span>
      </div>
      <div id="incident-progress" class="mt-3 d-none">
        <div class="progress" style="height:8px"><div class="progress-bar progress-bar-striped progress-bar-animated bg-warning" style="width:100%"></div></div>
        <p class="text-muted small mt-1">Processing…</p>
      </div>
      <div id="incident-result" class="mt-3 d-none"></div>
      <div class="mt-3 d-flex gap-2">
        <button class="btn btn-outline-secondary" onclick="goBack('incident')"><i class="bi bi-arrow-left me-1"></i>Back</button>
        <button class="btn btn-warning btn-lg px-5 text-dark" id="btn-upload-incident" disabled onclick="uploadFile('incident')">
          <i class="bi bi-upload me-2"></i>Upload &amp; Process Incidents
        </button>
      </div>
    </div>
  </div>
</div>

<!-- MACM Upload -->
<div id="upload-macm" class="upload-section {% if step != 3 %}d-none{% endif %}">
  <div class="card shadow border-0 upload-card macm-card">
    <div class="card-body p-4">
      <div class="d-flex align-items-center mb-3">
        <div class="icon-badge bg-success text-white rounded-3 p-3 me-3"><i class="bi bi-shield-check fs-3"></i></div>
        <div><h4 class="mb-0 fw-bold">Step 3 — Upload MACM File</h4><p class="text-muted mb-0 small">ServiceNow MACM export</p></div>
      </div>
      <div class="drop-zone" id="drop-macm" onclick="document.getElementById('file-macm').click()">
        <i class="bi bi-cloud-upload fs-1 text-success mb-2"></i>
        <p class="fw-semibold mb-1">Drag &amp; drop your MACM file here</p>
        <p class="text-muted small mb-0">or click to browse &nbsp;·&nbsp; Excel · CSV · PDF · Images</p>
        <input type="file" id="file-macm" class="d-none" accept=".xlsx,.xls,.csv,.pdf,.png,.jpg,.jpeg,.bmp,.tiff"/>
      </div>
      <div id="macm-file-info" class="mt-3 d-none">
        <i class="bi bi-file-check text-success me-2"></i>
        <span id="macm-filename" class="fw-semibold"></span>
        <span class="badge bg-success ms-2" id="macm-size"></span>
      </div>
      <div id="macm-progress" class="mt-3 d-none">
        <div class="progress" style="height:8px"><div class="progress-bar progress-bar-striped progress-bar-animated bg-success" style="width:100%"></div></div>
        <p class="text-muted small mt-1">Processing…</p>
      </div>
      <div id="macm-result" class="mt-3 d-none"></div>
      <!-- MACM row mapping -->
      <div class="mt-3 p-3 border rounded bg-light">
        <label class="form-label fw-semibold small mb-1">
          <i class="bi bi-table me-1 text-success"></i>Which row in your reference Excel should MACM data fill?
        </label>
        <select class="form-select form-select-sm" id="macm-row-label">
          <option value="operational - application ehancements">Operational – Application Enhancements (default)</option>
          <option value="operational - incident management">Operational – Incident Management</option>
          <option value="operational - request management">Operational – Request Management</option>
          <option value="operational">Operational (Total / Grand row)</option>
        </select>
        <div class="text-muted small mt-1">This maps MACM counts to the correct row when filling your Excel template.</div>
      </div>
      <div class="mt-3 d-flex gap-2">
        <button class="btn btn-outline-secondary" onclick="goBack('macm')"><i class="bi bi-arrow-left me-1"></i>Back</button>
        <button class="btn btn-success btn-lg px-5" id="btn-upload-macm" disabled onclick="uploadFile('macm')">
          <i class="bi bi-upload me-2"></i>Upload &amp; Process MACM
        </button>
      </div>
    </div>
  </div>
</div>

<!-- Team Efforts Upload (Step 4) -->
<div id="upload-team-efforts" class="upload-section {% if step != 4 %}d-none{% endif %}">
  <div class="card shadow border-0 upload-card" style="border-left:5px solid #0dcaf0!important">
    <div class="card-body p-4">
      <div class="d-flex align-items-center mb-3">
        <div class="icon-badge bg-info text-white rounded-3 p-3 me-3"><i class="bi bi-people-fill fs-3"></i></div>
        <div>
          <h4 class="mb-0 fw-bold">Step 4 — Upload Team Efforts Sheet</h4>
          <p class="text-muted mb-0 small">Your team efforts spreadsheet. RITM / Incident / MACM rows are filtered out — remaining entries count as <strong>Operational-JIRA&nbsp;/&nbsp;Deployments</strong>.</p>
        </div>
      </div>
      <div class="drop-zone" id="drop-team-efforts" onclick="document.getElementById('te-file').click()">
        <i class="bi bi-cloud-upload fs-1 text-info mb-2"></i>
        <p class="fw-semibold mb-1">Drag &amp; drop your Team Efforts files here</p>
        <p class="text-muted small mb-0">or click to browse &nbsp;·&nbsp; Excel · CSV &nbsp;·&nbsp; <strong>multiple files allowed</strong></p>
        <input type="file" id="te-file" class="d-none" accept=".xlsx,.xls,.csv" multiple/>
      </div>
      <div id="te-file-info" class="mt-3 d-none">
        <i class="bi bi-file-check text-success me-2"></i>
        <span id="te-filename" class="fw-semibold"></span>
      </div>
      <div id="te-status" class="mt-2 small"></div>
      <div class="mt-3 d-flex gap-2 flex-wrap">
        <button class="btn btn-outline-secondary" onclick="goBack('team_efforts')"><i class="bi bi-arrow-left me-1"></i>Back</button>
        <button id="btn-upload-te" class="btn btn-info btn-lg px-5 text-white" disabled onclick="uploadTeamEfforts()">
          <i class="bi bi-upload me-2"></i>Upload &amp; Continue
        </button>
        <button class="btn btn-outline-secondary btn-lg px-4" onclick="skipTeamEfforts()">
          <i class="bi bi-skip-forward me-1"></i>Skip
        </button>
      </div>
    </div>
  </div>
</div>

<!-- Reference Excel Upload (Step 5) -->
<div id="upload-reference" class="upload-section {% if step != 5 %}d-none{% endif %}">
  <div class="card shadow border-0 upload-card" style="border-left:5px solid #6f42c1!important">
    <div class="card-body p-4">
      <div class="d-flex align-items-center mb-3">
        <div class="icon-badge rounded-3 p-3 me-3 text-white" style="background:#6f42c1"><i class="bi bi-file-earmark-excel fs-3"></i></div>
        <div>
          <h4 class="mb-0 fw-bold">Step 5 — Upload Reference Excel</h4>
          <p class="text-muted mb-0 small">Your fixed-format monthly report template. Forecast values stay untouched — only Actual columns are filled.</p>
        </div>
      </div>
      <div class="drop-zone" id="drop-reference" onclick="document.getElementById('ref-file').click()">
        <i class="bi bi-file-earmark-excel fs-1 mb-2" style="color:#6f42c1"></i>
        <p class="fw-semibold mb-1">Drag &amp; drop your Reference Excel here</p>
        <p class="text-muted small mb-0">or click to browse &nbsp;·&nbsp; .xlsx or .xls only</p>
        <input type="file" id="ref-file" class="d-none" accept=".xlsx,.xls"/>
      </div>
      <div id="ref-file-info" class="mt-3 d-none">
        <i class="bi bi-file-check text-success me-2"></i>
        <span id="ref-filename" class="fw-semibold"></span>
        <span class="badge bg-success ms-2" id="ref-size"></span>
      </div>
      <div id="ref-status" class="mt-2 small"></div>
      <div class="mt-3 d-flex gap-2 flex-wrap">
        <button class="btn btn-outline-secondary" onclick="goBack('reference')"><i class="bi bi-arrow-left me-1"></i>Back</button>
        <button class="btn btn-lg px-5 text-white" id="btn-upload-ref" disabled onclick="uploadRef()" style="background:#6f42c1">
          <i class="bi bi-upload me-2"></i>Upload &amp; Continue
        </button>
        <button class="btn btn-outline-secondary btn-lg px-4" onclick="skipRef()">
          <i class="bi bi-skip-forward me-1"></i>Skip — generate new report
        </button>
        <a href="/download/reference-template" class="btn btn-outline-success btn-lg px-4" target="_blank">
          <i class="bi bi-download me-2"></i>Download Blank Template
        </a>
      </div>
    </div>
  </div>
</div>

<!-- Complete (Step 6) -->
<div id="upload-complete" class="upload-section {% if step < 6 %}d-none{% endif %}">
  <div class="card shadow border-0 p-4 p-md-5">
    <div class="text-center mb-4">
      <div class="success-animation"><i class="bi bi-check-circle-fill text-success" style="font-size:5rem"></i></div>
      <h3 class="fw-bold mt-3">All Files Processed!</h3>
      <p class="text-muted">Your ticket data is ready to view and export.</p>
    </div>
    <div class="row justify-content-center g-3 mb-4" id="summary-counts"></div>

    <!-- Month checkboxes + actions -->
    <div class="d-flex flex-column align-items-center gap-3 w-100">
      <div class="w-100" style="max-width:700px">
        <div class="d-flex align-items-center justify-content-between mb-2">
          <span class="fw-semibold"><i class="bi bi-calendar3 me-1 text-primary"></i>Select Month(s) to Fill</span>
          <div class="d-flex gap-2">
            <button class="btn btn-outline-secondary btn-sm" onclick="selectAllMonths(true)">Select All</button>
            <button class="btn btn-outline-secondary btn-sm" onclick="selectAllMonths(false)">Clear</button>
          </div>
        </div>
        <div id="monthCheckboxes" class="border rounded p-3 bg-light d-flex flex-wrap gap-2">
          <span class="text-muted small">Loading months…</span>
        </div>
        <p class="text-muted small mt-1 mb-0">If nothing selected → all months in your data will be filled.</p>
      </div>
      <div class="d-flex justify-content-center gap-3 flex-wrap">
        <a href="/dashboard" class="btn btn-primary btn-lg px-5"><i class="bi bi-bar-chart-fill me-2"></i>View Dashboard</a>
        <button onclick="downloadWithMonth()" class="btn btn-outline-success btn-lg px-4"><i class="bi bi-file-earmark-excel me-2"></i>Download Excel</button>
        <button onclick="downloadAndSend()" class="btn btn-success btn-lg px-4" id="btn-dl-send"><i class="bi bi-send me-2"></i>Download &amp; Send Email</button>
        <button onclick="sendEmailReport()" class="btn btn-outline-info btn-lg px-4" id="btn-send-email"><i class="bi bi-envelope me-2"></i>Send Email Only</button>
        <a href="/reset" class="btn btn-outline-secondary btn-lg px-4"><i class="bi bi-arrow-counterclockwise me-1"></i>Start Over</a>
      </div>
      <div id="email-status" class="small fw-semibold mt-1"></div>
    </div>

    <!-- Email Settings (collapsible) -->
    <div class="mt-4">
      <button class="btn btn-link p-0 text-muted small" type="button" data-bs-toggle="collapse" data-bs-target="#emailSettings">
        <i class="bi bi-gear me-1"></i>Email Settings
      </button>
      <div class="collapse mt-3" id="emailSettings">
        <div class="card card-body border bg-light">


          <div class="row g-3">
            <div class="col-md-5">
              <label class="form-label fw-semibold small">SMTP Relay Host</label>
              <input type="text" class="form-control form-control-sm" id="smtpHost" placeholder="mailrelay.cognizant.com or smtp.company.com"/>
            </div>
            <div class="col-md-2">
              <label class="form-label fw-semibold small">Port</label>
              <input type="number" class="form-control form-control-sm" id="smtpPort" placeholder="25"/>
            </div>
            <div class="col-md-5">
              <label class="form-label fw-semibold small">From Email <span class="text-muted">(sender address)</span></label>
              <input type="email" class="form-control form-control-sm" id="smtpUser" placeholder="your@cognizant.com"/>
            </div>
            <div class="col-md-6">
              <label class="form-label fw-semibold small">
                Password <span class="text-muted fw-normal">(leave blank for Internal Relay)</span>
                <a href="https://myaccount.google.com/apppasswords" target="_blank" class="ms-1 small text-muted">(Gmail App Password)</a>
              </label>
              <input type="password" class="form-control form-control-sm" id="smtpPassword" placeholder="Leave blank for relay / enter App Password for Gmail"/>
            </div>
            <div class="col-md-6">
              <label class="form-label fw-semibold small">Recipients <span class="text-muted">(one per line)</span></label>
              <textarea class="form-control form-control-sm" id="smtpRecipients" rows="2" placeholder="someone@cognizant.com&#10;another@email.com"></textarea>
            </div>
            <div class="col-12 d-flex gap-2 align-items-center flex-wrap">
              <button onclick="saveEmailConfig()" class="btn btn-primary btn-sm"><i class="bi bi-save me-1"></i>Save Settings</button>
              <button onclick="testEmailConnection()" class="btn btn-outline-secondary btn-sm" id="btn-test-email"><i class="bi bi-plug me-1"></i>Test Connection</button>
              <span id="email-cfg-status" class="small"></span>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>

</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script src="/assets/app.js"></script>
<script>initWizard({{ step | tojson }});</script>
{{ chatbot_widget | safe }}
</body></html>"""

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Ticket Dashboard</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet"/>
<link rel="stylesheet" href="/assets/style.css"/>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<style>
.month-tab { transition: all .15s; }
.month-tab.btn-primary { box-shadow: 0 2px 8px rgba(13,110,253,.35); }
.table-active td { background: #e7f1ff !important; }
</style>
</head>
<body class="bg-light">

<nav class="navbar navbar-dark" style="background:linear-gradient(135deg,#1a1a2e,#16213e,#0f3460)">
  <div class="container-fluid">
    <span class="navbar-brand fw-bold fs-5"><i class="bi bi-bar-chart-fill me-2 text-info"></i>Ticket Dashboard</span>
    <div class="d-flex gap-2">
      <a href="/download/report" class="btn btn-success btn-sm"><i class="bi bi-file-earmark-excel me-1"></i>Download Excel</a>
      <a href="/" class="btn btn-outline-light btn-sm"><i class="bi bi-upload me-1"></i>Upload More</a>
      <a href="/reset" class="btn btn-outline-warning btn-sm"><i class="bi bi-arrow-counterclockwise me-1"></i>Reset</a>
    </div>
  </div>
</nav>

<div class="container-fluid py-4 px-4">

<!-- ── Month Filter Bar ──────────────────────────────────────── -->
<div class="card border-0 shadow-sm mb-4">
  <div class="card-body py-3">
    <div class="d-flex align-items-center gap-3 flex-wrap">
      <div>
        <i class="bi bi-calendar3 me-1 text-primary"></i>
        <span class="fw-semibold small text-uppercase text-muted">Filter by Month</span>
      </div>
      <div id="monthTabs" class="d-flex flex-wrap gap-1"></div>
      <div class="ms-auto">
        <span class="badge bg-primary px-3 py-2 fs-6">
          <i class="bi bi-funnel me-1"></i><span id="filterLabel">All Months</span>
        </span>
      </div>
    </div>
  </div>
</div>

<!-- ── KPI Cards ─────────────────────────────────────────────── -->
<div class="row g-3 mb-4">
  <div class="col-sm-6 col-lg-3">
    <div class="card border-0 shadow-sm kpi-card"><div class="card-body d-flex align-items-center gap-3">
      <div class="kpi-icon bg-primary text-white rounded-3 p-3"><i class="bi bi-collection fs-3"></i></div>
      <div><div class="text-muted small fw-semibold text-uppercase">Grand Total</div><div class="fs-2 fw-bold" id="kpi-grand-total">—</div></div>
    </div></div>
  </div>
  <div class="col-sm-6 col-lg-3">
    <div class="card border-0 shadow-sm kpi-card"><div class="card-body d-flex align-items-center gap-3">
      <div class="kpi-icon bg-info text-white rounded-3 p-3"><i class="bi bi-list-check fs-3"></i></div>
      <div><div class="text-muted small fw-semibold text-uppercase">RITMs</div><div class="fs-2 fw-bold text-info" id="kpi-ritm">—</div></div>
    </div></div>
  </div>
  <div class="col-sm-6 col-lg-3">
    <div class="card border-0 shadow-sm kpi-card"><div class="card-body d-flex align-items-center gap-3">
      <div class="kpi-icon bg-warning text-dark rounded-3 p-3"><i class="bi bi-exclamation-triangle fs-3"></i></div>
      <div><div class="text-muted small fw-semibold text-uppercase">Incidents</div><div class="fs-2 fw-bold text-warning" id="kpi-incident">—</div></div>
    </div></div>
  </div>
  <div class="col-sm-6 col-lg-3">
    <div class="card border-0 shadow-sm kpi-card"><div class="card-body d-flex align-items-center gap-3">
      <div class="kpi-icon bg-success text-white rounded-3 p-3"><i class="bi bi-shield-check fs-3"></i></div>
      <div><div class="text-muted small fw-semibold text-uppercase">MACM</div><div class="fs-2 fw-bold text-success" id="kpi-macm">—</div></div>
    </div></div>
  </div>
</div>

<!-- ── Charts Row 1 ──────────────────────────────────────────── -->
<div class="row g-3 mb-4">
  <div class="col-lg-4">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header bg-white border-0 pt-3 pb-0"><h6 class="fw-bold mb-0"><i class="bi bi-pie-chart me-2 text-primary"></i>Ticket Type Split</h6></div>
      <div class="card-body d-flex justify-content-center align-items-center"><canvas id="chartDoughnut" style="max-height:280px"></canvas></div>
    </div>
  </div>
  <div class="col-lg-8">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header bg-white border-0 pt-3 pb-0">
        <h6 class="fw-bold mb-0"><i class="bi bi-graph-up me-2 text-success"></i>Monthly Ticket Trend
          <span class="badge bg-secondary fw-normal ms-2 small">All months — click a month tab above to highlight</span>
        </h6>
      </div>
      <div class="card-body"><canvas id="chartMonthly" style="max-height:280px"></canvas></div>
    </div>
  </div>
</div>

<!-- ── Monthly Summary Table ─────────────────────────────────── -->
<div class="card border-0 shadow-sm mb-4">
  <div class="card-header bg-white border-0 pt-3 pb-0 d-flex justify-content-between align-items-center">
    <h6 class="fw-bold mb-0"><i class="bi bi-calendar-range me-2 text-primary"></i>Monthly Breakdown — All Months</h6>
    <span class="text-muted small">Click any row to filter the entire dashboard to that month</span>
  </div>
  <div class="card-body p-0"><div class="table-responsive">
    <table class="table table-hover mb-0" id="monthlyTable">
      <thead class="table-dark"><tr>
        <th>Month</th>
        <th class="text-center">RITMs</th>
        <th class="text-center">Incidents</th>
        <th class="text-center">MACM</th>
        <th class="text-center">Total</th>
      </tr></thead>
      <tbody id="monthlyTbody" onclick="onMonthRowClick(event)"></tbody>
    </table>
  </div></div>
</div>

<!-- ── INC / RITM / MACM Monthly Summary ─────────────────────── -->
<div class="row g-3 mb-4 mt-2">
  <div class="col-xl-5 col-lg-12">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header bg-white border-0 pt-3 pb-2">
        <div class="d-flex align-items-center justify-content-between flex-wrap gap-2">
          <div class="d-flex align-items-center gap-2">
            <i class="bi bi-table text-primary"></i>
            <h6 class="fw-bold mb-0">INC / RITM / MACM Monthly Summary</h6>
          </div>
          <div class="d-flex align-items-center gap-2 flex-wrap">
            <label class="form-label mb-0 small fw-semibold text-muted">From</label>
            <select class="form-select form-select-sm" id="irmStartMonth" style="width:120px" onchange="filterIncRitmMacm()">
              <option value="">-- Start --</option>
            </select>
            <label class="form-label mb-0 small fw-semibold text-muted">To</label>
            <select class="form-select form-select-sm" id="irmEndMonth" style="width:120px" onchange="filterIncRitmMacm()">
              <option value="">-- End --</option>
            </select>
            <button class="btn btn-sm btn-outline-secondary px-2" onclick="resetIrmFilter()" title="Reset to all months">
              <i class="bi bi-x-circle me-1"></i>Reset
            </button>
          </div>
        </div>
      </div>
      <div class="card-body p-0">
        <div class="table-responsive">
          <table class="table table-bordered table-hover mb-0 small" id="incRitmMacmTable">
            <thead>
              <tr style="background:#1565c0;color:#fff">
                <th class="px-3" style="min-width:90px">Date</th>
                <th class="text-center">INC</th>
                <th class="text-center">RITM</th>
                <th class="text-center">MACM</th>
                <th class="text-center">MonthWiseTotal</th>
              </tr>
            </thead>
            <tbody id="incRitmMacmTbody">
              <tr><td colspan="5" class="text-center text-muted py-3">Loading…</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
  <div class="col-xl-7 col-lg-12">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header bg-white border-0 pt-3 pb-1 d-flex align-items-center gap-2">
        <i class="bi bi-bar-chart-fill text-primary"></i>
        <h6 class="fw-bold mb-0">INC/RITM/MACM</h6>
      </div>
      <div class="card-body"><canvas id="chartIncRitmMacm" style="max-height:320px"></canvas></div>
    </div>
  </div>
</div>

<!-- ── Work Types Report ─────────────────────────────────────── -->
<div class="row mt-4">
  <div class="col-12">
    <div class="card shadow-sm">
      <div class="card-header bg-primary text-white d-flex align-items-center gap-2">
        <i class="bi bi-pie-chart-fill fs-5"></i>
        <strong>Work Types Breakdown</strong>
        <span class="badge bg-light text-primary ms-auto" id="wtMonthBadge">All Months</span>
      </div>
      <div class="card-body">
        <div class="row g-3 align-items-center">
          <div class="col-lg-4 d-flex justify-content-center">
            <canvas id="chartWorkTypes" style="max-height:300px;max-width:300px"></canvas>
          </div>
          <div class="col-lg-8" id="workTypesTableWrap">
            <p class="text-muted text-center py-3">Loading&hellip;</p>
          </div>
        </div>
        <div class="row mt-3" id="wtMonthlyChartWrap" style="display:none">
          <div class="col-12">
            <canvas id="chartWorkTypesMonthly" style="max-height:260px"></canvas>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- ── Top 10 Slowest Tickets ──────────────────────────────── -->
<div class="row mt-4">
  <div class="col-12">
    <div class="card shadow-sm border-0">
      <div class="card-header text-white d-flex align-items-center gap-2"
           style="background:linear-gradient(135deg,#1a1a2e,#0f3460)">
        <i class="bi bi-hourglass-split fs-5"></i>
        <strong>Top 10 Slowest Tickets — Resolution Time</strong>
        <span class="badge bg-light text-dark ms-auto" style="font-size:.75rem">
          Duration = Opened &rarr; Closed / Resolved
        </span>
      </div>
      <div class="card-body p-3">
        <div class="row g-3">
          <!-- Incidents -->
          <div class="col-lg-6">
            <div class="d-flex align-items-center gap-2 mb-2">
              <span class="badge bg-warning text-dark fs-6 px-3 py-2">
                <i class="bi bi-exclamation-triangle me-1"></i>Incidents
              </span>
              <span class="text-muted small">Top 10 longest to resolve</span>
            </div>
            <div id="top10IncidentWrap">
              <p class="text-muted text-center py-3 small">Loading&hellip;</p>
            </div>
          </div>
          <!-- RITMs -->
          <div class="col-lg-6" style="border-left:1px solid #e9ecef">
            <div class="d-flex align-items-center gap-2 mb-2">
              <span class="badge bg-primary fs-6 px-3 py-2">
                <i class="bi bi-file-earmark-spreadsheet me-1"></i>RITMs
              </span>
              <span class="text-muted small">Top 10 longest to resolve</span>
            </div>
            <div id="top10RitmWrap">
              <p class="text-muted text-center py-3 small">Loading&hellip;</p>
            </div>
          </div>
        </div>
        <p class="text-muted small mb-0 mt-2">
          <i class="bi bi-info-circle me-1"></i>
          Duration is calculated from the ticket's opened/start date to its closed/resolved date.
          Tickets without both dates are excluded.
        </p>
      </div>
    </div>
  </div>
</div>

<!-- ── Duplicate Tickets Report ─────────────────────────────── -->
<div class="row mt-4">
  <div class="col-12">
    <div class="card shadow-sm">
      <div class="card-header bg-danger text-white py-2">
        <div class="d-flex align-items-center justify-content-between flex-wrap gap-2">
          <div class="d-flex align-items-center gap-2">
            <i class="bi bi-copy fs-5"></i>
            <strong>Duplicate Tickets Report</strong>
            <span class="badge bg-light text-danger">Same Short Description</span>
          </div>
          <div class="d-flex align-items-center gap-2 flex-wrap">
            <label class="form-label mb-0 small fw-semibold" style="opacity:.85">From</label>
            <select class="form-select form-select-sm text-dark" id="dupStartMonth" style="width:120px" onchange="filterDuplicates()">
              <option value="">-- Start --</option>
            </select>
            <label class="form-label mb-0 small fw-semibold" style="opacity:.85">To</label>
            <select class="form-select form-select-sm text-dark" id="dupEndMonth" style="width:120px" onchange="filterDuplicates()">
              <option value="">-- End --</option>
            </select>
            <button class="btn btn-sm btn-light text-danger px-2" onclick="resetDupFilter()" title="Reset to defaults">
              <i class="bi bi-x-circle me-1"></i>Reset
            </button>
            <div id="dupMonthChips" class="d-flex align-items-center gap-1 flex-wrap ms-1"></div>
          </div>
        </div>
      </div>
      <div class="card-body p-3" id="dupesWrap">
        <p class="text-muted text-center py-3">Loading&hellip;</p>
      </div>
    </div>
  </div>
</div>

</div><!-- /container -->

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script src="/assets/dashboard.js"></script>
<script>
// Sync secondary filter labels
const _origSetMonth = setMonth;
setMonth = function(m) {
  _origSetMonth(m);
  ["filterLabel2","filterLabel3"].forEach(id=>{
    const el=document.getElementById(id); if(el) el.textContent=m==="all"?"All Months":m;
  });
};
// Click on monthly table row to filter
function onMonthRowClick(e){
  const tr=e.target.closest("tr"); if(!tr)return;
  const mo=tr.cells[0]?.textContent.trim();
  if(mo && mo!=="TOTAL") setMonth(mo);
}
const RAW_DATA = { ritm: {{ ritm_data | safe }}, incident: {{ incident_data | safe }}, macm: {{ macm_data | safe }}, jira: {{ jira_data | safe }} };
initDashboard(RAW_DATA);
</script>
{{ chatbot_widget | safe }}
</body></html>"""

# ═══════════════════════════════════════════════════════════════
# 15. FLASK ROUTES
# ═══════════════════════════════════════════════════════════════

# ── Serve inline assets ───────────────────────────────────────
@app.route("/assets/style.css")
def asset_css():
    return CSS, 200, {"Content-Type": "text/css; charset=utf-8"}

@app.route("/assets/app.js")
def asset_app_js():
    return APP_JS, 200, {"Content-Type": "application/javascript; charset=utf-8"}

@app.route("/assets/dashboard.js")
def asset_dashboard_js():
    return DASHBOARD_JS, 200, {"Content-Type": "application/javascript; charset=utf-8"}

# ── Pages ─────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(INDEX_HTML, step=session.get("step", 1),
                                  chatbot_widget=CHATBOT_WIDGET)

@app.route("/reset")
def reset():
    _cleanup(); session.clear()
    return redirect(url_for("index"))

@app.route("/dashboard")
def dashboard():
    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    if not any([r, i, m]): return redirect(url_for("index"))
    te = _load("team_efforts") or {}
    return render_template_string(DASHBOARD_HTML,
        ritm_data=json.dumps(r), incident_data=json.dumps(i), macm_data=json.dumps(m),
        jira_data=json.dumps(te),
        chatbot_widget=CHATBOT_WIDGET)

# ── Upload ticket files ───────────────────────────────────────
@app.route("/upload/<ticket_type>", methods=["POST"])
def upload(ticket_type):
    tt = ticket_type.lower()
    if tt not in ("ritm", "incident", "macm"):
        return jsonify({"success": False, "error": "Unknown type"}), 400
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "Empty filename"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ALLOWED_EXT:
        return jsonify({"success": False, "error": f"Unsupported file type: .{ext}"}), 400

    try:
        _ensure_session()
        sd = UPLOAD_FOLDER / session["session_id"]; sd.mkdir(exist_ok=True)
        path = str(sd / f"{tt}_{secure_filename(f.filename)}")
        f.save(path)

        data = dispatch_file(path, tt.upper())
        _save(tt, data)

        if tt == "macm":
            ml = request.form.get("macm_label", "").strip()
            if ml:
                session["macm_label"] = ml

        step_map = {"ritm": 2, "incident": 3, "macm": 4}
        session["step"] = max(session.get("step", 1), step_map[tt])

        return jsonify({
            "success":    True,
            "ticket_type": tt.upper(),
            "total":       data.get("total", 0),
            "by_team":     data.get("by_team", {}),
            "by_assignee": data.get("by_assignee", {}),
            "by_month":    data.get("by_month", {}),
            "errors":      data.get("errors", []),
            "next_step":   session["step"]
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(exc) or "Server error processing file"}), 500

# ── Upload reference template ─────────────────────────────────
@app.route("/upload/reference", methods=["POST"])
def upload_reference():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "Must be .xlsx or .xls"}), 400
    try:
        _ensure_session()
        # Save per-session so each user keeps their own template
        sd = UPLOAD_FOLDER / session["session_id"]
        sd.mkdir(exist_ok=True)
        tmpl_path = sd / "reference_template.xlsx"
        f.save(str(tmpl_path))
        # Parse months from the uploaded template and cache in session
        tmpl_months = _get_template_months(str(tmpl_path))
        session["template_months"] = tmpl_months
        session["step"] = max(session.get("step", 1), 6)
        return jsonify({
            "success": True,
            "message": "Reference template uploaded. Proceeding to results.",
            "template_months": tmpl_months
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(exc) or "Server error processing reference file"}), 500

# ── Download blank reference template ────────────────────────
@app.route("/download/reference-template")
def download_reference_template():
    tmpl = REFERENCE_FOLDER / "reference.xlsx"
    create_reference_template(str(tmpl))
    return send_file(str(tmpl), as_attachment=True,
                     download_name="reference_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/upload/team_efforts", methods=["POST"])
def upload_team_efforts():
    files = request.files.getlist("file")
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({"success": False, "error": "No file selected"}), 400
    try:
        _ensure_session()
        sd = UPLOAD_FOLDER / session["session_id"]; sd.mkdir(exist_ok=True)
        merged = {"total": 0, "jira_count": 0, "ritm_count": 0,
                  "incident_count": 0, "macm_count": 0,
                  "by_month": {}, "records": [], "errors": []}
        for f in files:
            fname = f.filename or ""
            ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
            if ext not in {"xlsx", "xls", "csv"}:
                merged["errors"].append(f"Skipped {fname}: unsupported format .{ext}")
                continue
            path = str(sd / f"team_efforts_{secure_filename(fname)}")
            f.save(path)
            data = process_team_efforts(path)
            # Delete the raw uploaded file now that parsing is done.
            try:
                os.remove(path)
            except Exception:
                pass
            merged["total"]          += data.get("total", 0)
            merged["jira_count"]     += data.get("jira_count", 0)
            merged["ritm_count"]     += data.get("ritm_count", 0)
            merged["incident_count"] += data.get("incident_count", 0)
            merged["macm_count"]     += data.get("macm_count", 0)
            merged["records"].extend(data.get("records", []))
            merged["errors"].extend(data.get("errors", []))
            for k, v in data.get("by_month", {}).items():
                merged["by_month"][k] = merged["by_month"].get(k, 0) + v
        _save("team_efforts", merged)
        session["step"] = max(session.get("step", 1), 5)
        return jsonify({
            "success":        True,
            "total":          merged["total"],
            "jira_count":     merged["jira_count"],
            "ritm_count":     merged["ritm_count"],
            "incident_count": merged["incident_count"],
            "macm_count":     merged["macm_count"],
            "errors":         merged["errors"]
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(exc) or "Server error processing team efforts"}), 500

# ── Download report ───────────────────────────────────────────
@app.route("/download/report")
def download_report():
    _ensure_session()
    try:
        r = _load("ritm"); i = _load("incident"); m = _load("macm")
        if not r and not i and not m:
            return "<h3>No data found. Please go back and upload your files first.</h3>" \
                   "<a href='/'>Go back</a>", 400

        months_param = request.args.get("months") or ""
        selected_months = [mo.strip() for mo in months_param.split(",") if mo.strip()] or None
        month_name = datetime.now().strftime("%B")
        name = f"SupportOCI_30-60-90 Forecast and Actuals_v1.1.2_{month_name}.xlsx"
        out  = str(OUTPUT_FOLDER / name)
        tmpl = _get_session_template()

        macm_label  = session.get("macm_label")
        te          = _load("team_efforts")
        if tmpl:
            try:
                res = fill_reference_excel(tmpl, out, r, i, m, selected_months,
                                           macm_label=macm_label, team_efforts=te)
                if res.get("success"):
                    return send_file(out, as_attachment=True, download_name=name,
                                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception:
                pass  # fall through to standalone report

        generate_standalone_report(out, r, i, m, te, selected_months=selected_months)
        return send_file(out, as_attachment=True, download_name=name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return f"<h3>Error generating report: {e}</h3><a href='/'>Go back</a>", 500

# ── API helpers ───────────────────────────────────────────────
@app.route("/api/rawdata")
def api_rawdata():
    return jsonify({
        "ritm":     _load("ritm"),
        "incident": _load("incident"),
        "macm":     _load("macm"),
    })

@app.route("/api/months")
def api_months():
    """Return sorted unique months in 'Mar-2026' format from all uploaded ticket data."""
    _ensure_session()
    try:
        r = _load("ritm"); i = _load("incident"); m = _load("macm")
        all_keys = set()
        for d in [r, i, m]:
            try:
                all_keys.update(_monthly_counts(d).keys())
            except Exception:
                pass
        # Filter out non-month keys like 'NaT'
        valid = [k for k in all_keys if re.match(r"^[A-Za-z]{3}-\d{4}$", k)]
        months = sorted(valid, key=lambda k: (
            int(k.split("-")[1]),
            MONTH_ABBR.index(k.split("-")[0]) if k.split("-")[0] in MONTH_ABBR else 0
        ))
        return jsonify({"months": months})
    except Exception as e:
        return jsonify({"months": [], "error": str(e)})

@app.route("/api/template-months")
def api_template_months():
    """Return the months detected in the uploaded reference Excel template."""
    _ensure_session()
    # Use cached value from session if available
    cached = session.get("template_months")
    if cached is not None:
        return jsonify({"months": cached, "has_template": True})
    tmpl = _get_session_template()
    if not tmpl:
        return jsonify({"months": [], "has_template": False})
    months = _get_template_months(tmpl)
    session["template_months"] = months
    return jsonify({"months": months, "has_template": bool(months)})


@app.route("/api/summary")
def api_summary():
    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    return jsonify({
        "ritm_total":     r.get("total", 0) if r else None,
        "incident_total": i.get("total", 0) if i else None,
        "macm_total":     m.get("total", 0) if m else None,
        "step":           session.get("step", 1)
    })

@app.route("/api/work-types", methods=["GET"])
def api_work_types():
    month = request.args.get("month", "all").strip()
    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    te = _load("team_efforts")

    # Convert "Jan-2024" → "2024-01" prefix for record-level filtering
    month_prefix = None
    if month and month != "all":
        parts = month.strip().split("-")
        if len(parts) == 2 and parts[0] in MONTH_ABBR and parts[1].isdigit():
            month_prefix = f"{parts[1]}-{str(MONTH_ABBR.index(parts[0]) + 1).zfill(2)}"

    def mc(data, mo):
        if not data: return 0
        if mo == "all": return data.get("total", 0)
        monthly = _monthly_counts(data)      # keys are Mon-YYYY
        if mo in monthly: return monthly[mo]
        # Convert YYYY-MM → Mon-YYYY and retry
        try:
            parts = mo.split("-")
            if len(parts) == 2 and len(parts[0]) == 4 and parts[0].isdigit():
                abbr = f"{MONTH_ABBR[int(parts[1])-1]}-{parts[0]}"
                return monthly.get(abbr, 0)
        except Exception: pass
        return 0

    db_patch = _count_db_patching(r or {}, month_prefix)

    jira = 0
    if te:
        if month == "all":
            jira = te.get("jira_count", 0)
        else:
            bm = te.get("by_month", {})
            # te["by_month"] stores Mon-YYYY keys; convert YYYY-MM if needed
            if month in bm:
                jira = bm[month]
            else:
                try:
                    parts = month.split("-")
                    if len(parts) == 2 and len(parts[0]) == 4 and parts[0].isdigit():
                        abbr = f"{MONTH_ABBR[int(parts[1])-1]}-{parts[0]}"
                        jira = bm.get(abbr, 0)
                    else:
                        jira = bm.get(month, 0)
                except Exception:
                    jira = 0

    work_types = [
        {"label": "Operational - Request Management",
         "description": "RITM tickets raised via ServiceNow",
         "count": mc(r, month), "color": "#0d6efd"},
        {"label": "Operational - Incident Management",
         "description": "Incident tickets raised via ServiceNow",
         "count": mc(i, month), "color": "#ffc107"},
        {"label": "Operational - Application Enhancements",
         "description": "MACM tickets raised via ServiceNow",
         "count": mc(m, month), "color": "#198754"},
        {"label": "Application Enhancements",
         "description": ("Optimizing server and application security/reliability by analyzing and "
                         "applying the latest required application and database patches on "
                         "monthly and quarterly basis."),
         "count": db_patch, "color": "#6f42c1"},
    ]
    if te:
        work_types.append({
            "label": "Operational-JIRA",
            "description": "Deployments",
            "count": jira, "color": "#fd7e14"
        })
    return jsonify({"work_types": work_types, "month": month, "has_team_efforts": bool(te)})

@app.route("/api/email-config", methods=["GET","POST"])
def api_email_config():
    if request.method == "POST":
        data = request.get_json(force=True) or {}
        cfg  = load_email_config()
        cfg.update({k: data[k] for k in data if k in DEFAULT_EMAIL_CONFIG})
        # Strip whitespace from credentials to avoid leading/trailing space issues
        for field in ("smtp_host", "smtp_user", "smtp_password"):
            if field in cfg and isinstance(cfg[field], str):
                cfg[field] = cfg[field].strip()
        # recipients may come as a newline-separated string from the textarea
        if "recipients_raw" in data:
            cfg["recipients"] = [r.strip() for r in data["recipients_raw"].splitlines() if r.strip()]
        save_email_config(cfg)
        return jsonify({"success": True})
    cfg = load_email_config()
    # Don't send password to frontend
    safe = dict(cfg); safe["smtp_password"] = "" if cfg.get("smtp_password") else ""
    safe["recipients_raw"] = "\n".join(cfg.get("recipients", []))
    return jsonify(safe)

@app.route("/api/send-email", methods=["POST"])
def api_send_email():
    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    if not any([r, i, m]):
        return jsonify({"success": False, "error": "No ticket data. Upload files first."}), 400
    data            = request.get_json(force=True) or {}
    selected_months = data.get("months") or None
    if isinstance(selected_months, list):
        selected_months = [m.strip() for m in selected_months if m.strip()] or None

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"Ticket_Report_{ts}.xlsx"
    out  = str(OUTPUT_FOLDER / name)
    tmpl = _get_session_template()

    te         = _load("team_efforts")
    macm_label = session.get("macm_label")
    if tmpl:
        res = fill_reference_excel(tmpl, out, r, i, m, selected_months,
                                   macm_label=macm_label, team_efforts=te)
        if not res["success"]:
            generate_standalone_report(out, r, i, m, te, selected_months=selected_months)
    else:
        generate_standalone_report(out, r, i, m, te, selected_months=selected_months)

    label = ", ".join(selected_months) if selected_months else "All Months"

    # Run the SMTP send in a background thread so the browser gets an immediate
    # response.  The UI polls /api/email-status/<job_id> every 2 s for the result.
    job_id = str(uuid.uuid4())
    _email_jobs[job_id] = {"status": "sending"}

    def _do_send(path, lbl, jid):
        result = send_email_report(path, lbl)
        result["status"] = "done"
        _email_jobs[jid] = result

    threading.Thread(target=_do_send, args=(out, label, job_id), daemon=True).start()
    return jsonify({"status": "sending", "job_id": job_id,
                    "message": "Sending email in background…"})

@app.route("/api/email-status/<job_id>")
def api_email_status(job_id):
    job = _email_jobs.get(job_id)
    if job is None:
        return jsonify({"status": "unknown"}), 404
    return jsonify(job)

@app.route("/api/test-email", methods=["POST"])
def api_test_email():
    """Send a short test email to verify SMTP settings work."""
    cfg = load_email_config()
    recipients = [r.strip() for r in cfg.get("recipients", []) if r.strip()]
    if not recipients:
        return jsonify({"success": False, "error": "No recipients configured. Add at least one recipient."})
    use_auth = bool(cfg.get("smtp_user", "").strip() and cfg.get("smtp_password", "").strip())
    job_id = str(uuid.uuid4())
    _email_jobs[job_id] = {"status": "sending"}

    def _do_test(jid):
        try:
            sender_addr = cfg.get("smtp_user", "").strip() or cfg.get("sender_email", "noreply@company.com")
            msg = MIMEMultipart()
            msg["From"]    = f"{cfg['sender_name']} <{sender_addr}>"
            msg["To"]      = ", ".join(recipients)
            msg["Subject"] = "Test — ServiceNow Ticket Counter SMTP Check"
            msg.attach(MIMEText(
                "This is a test email from the ServiceNow Ticket Counter app.\n"
                "If you received this, your SMTP settings are correct!", "plain"))
            context = ssl.create_default_context()
            with smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"]), timeout=15) as server:
                if cfg.get("use_tls", False) and use_auth:
                    server.starttls(context=context)
                if use_auth:
                    server.login(cfg["smtp_user"].strip(), cfg["smtp_password"].strip())
                server.sendmail(sender_addr, recipients, msg.as_string())
            mode = "relay (no auth)" if not use_auth else "authenticated"
            _email_jobs[jid] = {"status": "done", "success": True,
                                 "message": f"Test email sent via {mode} to {', '.join(recipients)}"}
        except TimeoutError:
            _email_jobs[jid] = {"status": "done", "success": False,
                                 "error": f"Timed out — {cfg['smtp_host']}:{cfg['smtp_port']} unreachable. Check relay server address."}
        except smtplib.SMTPAuthenticationError:
            _email_jobs[jid] = {"status": "done", "success": False,
                                 "error": "Authentication failed — wrong email or password."}
        except smtplib.SMTPServerDisconnected:
            _email_jobs[jid] = {"status": "done", "success": False,
                                 "error": f"Server {cfg['smtp_host']} closed the connection. Check the relay host address."}
        except Exception as e:
            _email_jobs[jid] = {"status": "done", "success": False, "error": str(e)}

    threading.Thread(target=_do_test, args=(job_id,), daemon=True).start()
    return jsonify({"status": "sending", "job_id": job_id})

# ── Chatbot API ───────────────────────────────────────────────
@app.route("/api/chat", methods=["POST"])
def api_chat():
    _ensure_session()
    data    = request.get_json(force=True) or {}
    message = data.get("message", "").strip()
    history = data.get("history", [])
    if not message:
        return jsonify({"reply": "Please type a message."})

    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    all_data = {"ritm": r or {}, "incident": i or {}, "macm": m or {}}

    # ── Top-10 slow tickets query detection ───────────────────
    msg_l = message.lower()
    _top10_kw = ["top 10", "top10", "slowest", "longest", "long time",
                 "slow ticket", "most time", "duration", "taking long",
                 "resolution time", "overdue", "breach"]
    if any(kw in msg_l for kw in _top10_kw):
        reply = _answer_top10_slow(r or {}, i or {})
        return jsonify({"reply": reply, "tickets_count": 0})

    # 1. Try exact/partial ticket number match
    tickets_found = []
    ticket_pat = re.search(r"(?:RITM|INC|MACM|TASK|REQ)\d+", message, re.IGNORECASE)
    if ticket_pat:
        tickets_found = _search_ticket(ticket_pat.group(0), all_data)

    # 2. Keyword search if no ticket number found and message is long enough
    if not tickets_found and len(message) > 3:
        tickets_found = _search_by_keyword(message, all_data, limit=3)

    data_context = _build_data_context(all_data)
    reply = _generate_chat_response(message, tickets_found, data_context, all_data, history)
    return jsonify({"reply": reply, "tickets_count": len(tickets_found)})

@app.route("/api/chatbot-config", methods=["GET", "POST"])
def api_chatbot_config():
    if request.method == "POST":
        data = request.get_json(force=True) or {}
        cfg  = load_chatbot_config()
        if data.get("api_key", "").strip():
            cfg["api_key"] = data["api_key"].strip()
        if data.get("model", "").strip():
            cfg["model"] = data["model"].strip()
        save_chatbot_config(cfg)
        return jsonify({"success": True})
    cfg = load_chatbot_config()
    return jsonify({
        "has_key": bool(cfg.get("api_key")),
        "model":   cfg.get("model", "claude-haiku-4-5-20251001"),
    })

@app.route("/api/duplicates", methods=["GET"])
def api_duplicates():
    month = request.args.get("month", "all").strip()
    r = _load("ritm"); i = _load("incident"); m = _load("macm")
    all_data = {"ritm": r, "incident": i, "macm": m}
    dupes, monthly_summary = _find_duplicates(all_data, month=month)
    return jsonify({"duplicates": dupes, "total_groups": len(dupes),
                    "monthly_summary": monthly_summary, "month": month})

@app.route("/api/top10-slow", methods=["GET"])
def api_top10_slow():
    _ensure_session()
    r = _load("ritm"); i = _load("incident")

    def _top10_for(src, ttype):
        recs = (src.get("top10_slow") or []) if src else []
        # keep only matching type and sort by duration desc
        recs = [rec for rec in recs if rec.get("ticket_type","").upper() == ttype.upper()]
        recs.sort(key=lambda x: x.get("duration_days", 0), reverse=True)
        return recs[:10]

    ritm_top10     = _top10_for(r, "RITM")
    incident_top10 = _top10_for(i, "INCIDENT")
    combined       = _get_top10_slow(r or {}, i or {})
    return jsonify({
        "ritm":     ritm_top10,
        "incident": incident_top10,
        "combined": combined,
        "count":    len(combined)
    })

# ═══════════════════════════════════════════════════════════════
# 16. MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Recreate default template if it has old short-name labels (RITM/Incident/MACM)
    _tmpl_path = str(REFERENCE_FOLDER / "reference.xlsx")
    if os.path.exists(_tmpl_path):
        try:
            _wb_chk = openpyxl.load_workbook(_tmpl_path, read_only=True, data_only=True)
            _old_labels = {"ritm", "incident", "macm"}
            _needs_regen = False
            for _ws_chk in _wb_chk.worksheets:
                for _row in _ws_chk.iter_rows(min_row=4, max_row=4, min_col=1, max_col=1, values_only=True):
                    if str(_row[0] or "").strip().lower() in _old_labels:
                        _needs_regen = True
                        break
                if _needs_regen:
                    break
            _wb_chk.close()
            if _needs_regen:
                os.remove(_tmpl_path)
                print("Auto-regenerating reference template with updated labels…")
        except Exception:
            pass
    create_reference_template(_tmpl_path)

    # Auto-open browser
    def _open():
        import time; time.sleep(1.5)
        webbrowser.open("http://localhost:2020")
    threading.Thread(target=_open, daemon=True).start()

    ocr_method = ("PowerShell/Windows OCR" if POWERSHELL_OK
                  else "tesserocr" if TESSERACT_OK
                  else "NONE (images unsupported)")
    port = int(os.environ.get("PORT", 2020))
    print("=" * 58)
    print("  ServiceNow Ticket Counter")
    print(f"  URL: http://localhost:{port}")
    print(f"  Image OCR: {ocr_method}")
    print("  Press Ctrl+C to stop")
    print("=" * 58)
    app.run(debug=False, host="0.0.0.0", port=port)